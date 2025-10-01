"""
Sign Estimation Application
A modern Python web application for sign manufacturing cost estimation and project management.
"""

import os
import sys
import socket
import dash
from dash import html, dcc, Input, Output, State, callback_context, dash_table
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
import plotly.express as px
import dash_cytoscape as cyto
import pandas as pd
import sqlite3
from datetime import datetime, timezone
import base64
import io
import json
from pathlib import Path
from dash.exceptions import PreventUpdate

# Lightweight image type detector (avoid imghdr dependency warnings)
def _detect_image_type(raw: bytes, filename: str) -> str | None:
    head = raw[:16]
    lower_name = filename.lower()
    if head.startswith(b"\x89PNG\r\n\x1a\n"):
        return 'png'
    if head.startswith(b"\xff\xd8"):
        return 'jpg'
    if head.startswith(b"GIF87a") or head.startswith(b"GIF89a"):
        return 'gif'
    # SVG (either by extension or xml/svg tag early)
    if lower_name.endswith('.svg') or b'<svg' in raw[:200].lower():
        return 'svg'
    return None

# ---------------- Application Initialization (restored after consolidation) ---------------- #
from config import DATABASE_PATH as _BASE_DB_PATH, APP_HOST, APP_PORT, DASH_DEBUG, ensure_backup_dir, AUTO_BACKUP_INTERVAL_SEC, BACKUP_DIR, ONEDRIVE_SYNC_DIR, ONEDRIVE_AUTOSYNC_SEC  # type: ignore
DATABASE_PATH = os.getenv("SIGN_APP_DB", _BASE_DB_PATH)

# Make utils importable
sys.path.append(str(Path(__file__).parent / 'utils'))
try:
    from utils.database import DatabaseManager  # type: ignore
    from utils.calculations import CostCalculator, compute_unit_price, compute_install_cost  # type: ignore
    from utils.onedrive import OneDriveManager  # type: ignore
except Exception as e:  # Fallback minimal stubs to keep module importable
    print(f"[startup][warn] Failed importing utils modules: {e}")
    class DatabaseManager:  # type: ignore
        def __init__(self, db_path): self.db_path = db_path
        def init_database(self): pass
        def create_pricing_profile(self,*a,**kw): return 0
        def assign_pricing_profile_to_project(self,*a,**kw): return None
        def add_note(self,*a,**kw): return None
        def list_notes(self,*a,**kw): return []
    class CostCalculator:  # type: ignore
        def __init__(self, db_path): pass
    def compute_unit_price(row_dict, price_mode): return float(row_dict.get('unit_price') or 0)
    def compute_install_cost(*a,**kw): return 0.0
    class OneDriveManager:  # type: ignore
        def __init__(self, local_path): pass
        def sync_database(self): return False, 'onedrive disabled'

db_manager = DatabaseManager(DATABASE_PATH)
try:
    db_manager.init_database()
except Exception as e:
    print(f"[startup][error] init_database: {e}")
cost_calculator = CostCalculator(DATABASE_PATH)
onedrive_manager = OneDriveManager(Path.cwd())

def ensure_extended_schema():
    try:
        conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
        cur.execute("PRAGMA table_info('sign_types')"); cols = {r[1] for r in cur.fetchall()}
        if 'material_alt' not in cols:
            try: cur.execute("ALTER TABLE sign_types ADD COLUMN material_alt TEXT")
            except Exception: pass
        if 'material_multiplier' not in cols:
            try: cur.execute("ALTER TABLE sign_types ADD COLUMN material_multiplier REAL DEFAULT 0")
            except Exception: pass
        if 'install_type' not in cols:
            try: cur.execute("ALTER TABLE sign_types ADD COLUMN install_type TEXT")
            except Exception: pass
        if 'install_time_hours' not in cols:
            try: cur.execute("ALTER TABLE sign_types ADD COLUMN install_time_hours REAL DEFAULT 0")
            except Exception: pass
        if 'per_sign_install_rate' not in cols:
            try: cur.execute("ALTER TABLE sign_types ADD COLUMN per_sign_install_rate REAL DEFAULT 0")
            except Exception: pass
        cur.execute("PRAGMA table_info('building_signs')"); bcols = {r[1] for r in cur.fetchall()}
        if 'custom_price' not in bcols:
            try: cur.execute("ALTER TABLE building_signs ADD COLUMN custom_price REAL")
            except Exception: pass
        conn.commit(); conn.close()
    except Exception as e:
        print(f"[schema][warn] {e}")

ensure_extended_schema()

# Dash app
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP], suppress_callback_exceptions=True)
app.title = "Sign Package Estimator"

# Runtime cross-platform interpreter path sanity check
def _interpreter_sanity():
    bad_markers = ['/Users/', '/miniconda3/', '/bin/python']
    # Some libs set PYTHONEXECUTABLE (PyInstaller, etc.)
    suspect = os.environ.get('PYTHONEXECUTABLE') or os.environ.get('PYTHON_EXECUTABLE')
    if suspect and os.name == 'nt':
        lower = suspect.replace('\\','/').lower()
        if any(m in lower for m in ['/users/', '/miniconda3/']):
            print(f"[startup][warn] Ignoring non-Windows interpreter reference: {suspect}")
    # Defensive: ensure sys.executable exists
    if not os.path.exists(sys.executable):
        print(f"[startup][warn] sys.executable missing ({sys.executable}); continuing but environment may be broken.")
    # Cross-platform venv mismatch detection (e.g., pyvenv.cfg referencing mac path on Windows)
    try:
        venv_cfg = Path(sys.prefix) / 'pyvenv.cfg'
        mismatch = False
        origin = None
        if venv_cfg.exists():
            text = venv_cfg.read_text(errors='ignore')
            if os.name == 'nt' and '\nhome = /users/' in text.lower():
                mismatch = True; origin = 'macOS'
            if os.name != 'nt' and '\\python.exe' in text.lower():
                mismatch = True; origin = 'Windows'
        if mismatch:
            os.environ['SIGN_APP_ENV_MISMATCH'] = f"venv-origin:{origin}"  # used by banner callback
            print(f"[startup][env][warn] Detected cross-platform virtualenv mismatch (origin {origin}). Recommend recreating venv on this platform.")
    except Exception as e:
        print(f"[startup][env][warn] mismatch detection failed: {e}")

_interpreter_sanity()

# Removed _cairosvg_probe (previously set SIGN_APP_SVG_STATUS) as banners were eliminated.

# Consolidated role/tab globals
ROLE_OPTIONS = [
    {'label': 'Admin', 'value': 'admin'},
    {'label': 'Estimator', 'value': 'estimator'},
    {'label': 'Sales', 'value': 'sales'},
    {'label': 'Viewer', 'value': 'viewer'}
]
ROLE_RANK = {'admin':4,'estimator':3,'sales':2,'viewer':1}
TAB_DEFS = [
    ('projects-tab','Projects','viewer'),
    ('signs-tab','Sign Types','estimator'),
    ('groups-tab','Sign Groups','estimator'),
    ('building-tab','Building View','viewer'),
    ('estimates-tab','Estimates','viewer'),
    ('import-tab','Import Data','viewer'),
    ('tab_profiles','Pricing Profiles','admin'),
    ('tab_snapshots','Snapshots','estimator'),
    ('tab_templates','Bid Templates','estimator'),
    ('tab_tags','Tags','estimator'),
    ('tab_notes','Notes','estimator')
]
def build_tabs_for_role(role: str):
    rank = ROLE_RANK.get(role or 'viewer',1)
    return [dbc.Tab(label=lbl, tab_id=tid) for tid,lbl,min_role in TAB_DEFS if ROLE_RANK.get(min_role,1) <= rank]

# Add utils to path
@app.callback(
    Output('note-add-feedback','children'),
    Input('add-note-btn','n_clicks'),
    State('note-entity-type','value'),
    State('note-entity-id','value'),
    State('note-text','value'),
    State('note-include','value'),
    prevent_initial_call=True
)
def add_note_cb(n, etype, ent_id, text, include_values):
    if not n:
        raise PreventUpdate
    if not text or not ent_id:
        return dbc.Alert('Entity and text required', color='danger')
    try:
        # For sign_type we need name -> id lookup in notes table design uses entity_id; we will allow sign_types by name mapping here
        if etype == 'sign_type' and not str(ent_id).isdigit():
            conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
            cur.execute('SELECT id FROM sign_types WHERE lower(name)=lower(?)',(str(ent_id),))
            row = cur.fetchone(); conn.close()
            if not row:
                return dbc.Alert('Sign type not found', color='danger')
            ent_key = row[0]
        else:
            ent_key = int(ent_id)
        inc = 'inc' in (include_values or [])
        db_manager.add_note(etype, ent_key, text.strip(), include_in_export=inc)
        return dbc.Alert('Note added', color='success', dismissable=True)
    except Exception as e:
        return dbc.Alert(f'Error: {e}', color='danger')

@app.callback(
    Output('notes-table','data'),
    Input('load-notes-btn','n_clicks'),
    State('list-note-entity-type','value'),
    State('list-note-entity-id','value'),
    prevent_initial_call=True
)
def load_notes_cb(n, etype, ent_id):
    if not n:
        raise PreventUpdate
    if not ent_id:
        return []
    try:
        if etype == 'sign_type' and not str(ent_id).isdigit():
            conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
            cur.execute('SELECT id FROM sign_types WHERE lower(name)=lower(?)',(str(ent_id),))
            row = cur.fetchone(); conn.close()
            if not row:
                return []
            ent_key = row[0]
        else:
            ent_key = int(ent_id)
        rows = db_manager.list_notes(etype, ent_key, export_only=False)
        # rows: id, note, include_in_export?, created_at (depending on select)
        data = []
        for r in rows:
            if len(r) == 4:
                nid, note, include_flag, created = r
                data.append({'id':nid,'note':note,'include_in_export':'Yes' if include_flag else '','created_at':created})
            else:
                # export_only path returns 3 columns
                nid, note, created = r
                data.append({'id':nid,'note':note,'include_in_export':'(export)','created_at':created})
        return data
    except Exception:
        return []

@app.callback(
    Output('toggle-note-feedback','children'),
    Input('toggle-note-btn','n_clicks'),
    State('toggle-note-id','value'),
    prevent_initial_call=True
)
def toggle_note_include(n, note_id):
    if not n:
        raise PreventUpdate
    if not note_id:
        return dbc.Alert('Note ID required', color='danger')
    try:
        conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
        cur.execute('SELECT include_in_export FROM notes WHERE id=?',(int(note_id),))
        row = cur.fetchone()
        if not row:
            conn.close(); return dbc.Alert('Note not found', color='danger')
        new_val = 0 if row[0] else 1
        cur.execute('UPDATE notes SET include_in_export=? WHERE id=?',(new_val, int(note_id)))
        conn.commit(); conn.close()
        return dbc.Alert('Include flag toggled', color='success', dismissable=True)
    except Exception as e:
        return dbc.Alert(f'Error: {e}', color='danger')

# ------------------- Dropdown Options Refresh ------------------- #
from dash.dependencies import ALL
@app.callback(
    Output('snap-project-id','options'),
    Output('pp-project-id','options'),
    Output('pp-profile-id','options'),
    Output('pp-default-id','options'),
    Output('template-id-item','options'),
    Output('apply-template-id','options'),
    Input({'kind':ALL,'action':ALL}, 'n_clicks'),
    Input('global-dropdown-refresh','n_intervals'),
    State('template-order-mode','value'),
    prevent_initial_call=False
)
def refresh_dropdown_options(_pattern_clicks, _interval, template_order_mode):
    try:
        conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
        cur.execute('SELECT id, name FROM projects ORDER BY name')
        projects = [{'label':r[1], 'value': r[0]} for r in cur.fetchall()]
        cur.execute('SELECT id, name FROM pricing_profiles ORDER BY name')
        profiles = [{'label':r[1], 'value': r[0]} for r in cur.fetchall()]
        if template_order_mode == 'alpha':
            cur.execute('SELECT id, name FROM bid_templates ORDER BY lower(name) ASC')
        else:
            cur.execute('SELECT id, name FROM bid_templates ORDER BY created_at DESC')
        templates = [{'label':r[1], 'value': r[0]} for r in cur.fetchall()]
        # apply-template-id uses same templates list
        conn.close()
        return projects, projects, profiles, profiles, templates, templates
    except Exception as e:
        print(f"[dropdown-refresh][error] {e}")
        empty = []
        return empty, empty, empty, empty, empty, empty

# ------------------- Pricing Profiles Callbacks ------------------- #
@app.callback(
    Output('pp-create-feedback','children'),
    Output('pp-table','data'),
    Input({'kind':'pp','action':ALL}, 'n_clicks'),
    State('pp-name','value'),
    State('pp-tax','value'),
    State('pp-install','value'),
    State('pp-margin','value'),
    State('pp-default','value'),
    prevent_initial_call=True
)
def handle_pricing_profiles(pp_clicks, name, tax, install, margin, default_values):
    # Determine which button fired
    triggered_actions = []
    if callback_context.triggered:
        for t in callback_context.triggered:
            try:
                tid = eval(t['prop_id'].split('.')[0])  # pattern IDs are serialized as dict strings
                if isinstance(tid, dict) and tid.get('kind')=='pp':
                    triggered_actions.append(tid.get('action'))
            except Exception:
                continue
    msg = dash.no_update
    if 'create' in triggered_actions:
        if not name:
            return dbc.Alert('Profile name required', color='danger'), dash.no_update
        try:
            pid = db_manager.create_pricing_profile(
                name.strip(),
                float(tax or 0)/100.0,
                float(install or 0)/100.0,
                float(margin or 1) or 1.0,
                is_default=('d' in (default_values or []))
            )
            msg = dbc.Alert(f'Created profile (ID {pid})', color='success', dismissable=True)
        except Exception as e:
            msg = dbc.Alert(f'Error: {e}', color='danger')
    # Load all profiles
    conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
    cur.execute('SELECT id, name, sales_tax_rate, installation_rate, margin_multiplier, is_default FROM pricing_profiles ORDER BY id DESC')
    rows = [
        {
            'id':r[0], 'name':r[1],
            'sales_tax_rate': f"{(r[2] or 0)*100:.2f}%",
            'installation_rate': f"{(r[3] or 0)*100:.2f}%",
            'margin_multiplier': f"{r[4]:.3f}",
            'is_default': 'Yes' if r[5] else ''
        } for r in cur.fetchall()
    ]
    conn.close()
    return msg, rows

@app.callback(
    Output('pp-assign-feedback','children'),
    Input('pp-assign-btn','n_clicks'),
    State('pp-project-id','value'),
    State('pp-profile-id','value'),
    prevent_initial_call=True
)
def assign_pricing_profile(n, project_id, profile_id):
    if not n:
        raise PreventUpdate
    if not (project_id and profile_id):
        return dbc.Alert('Project ID and Profile ID required', color='danger')
    try:
        db_manager.assign_pricing_profile_to_project(int(project_id), int(profile_id))
        return dbc.Alert('Profile assigned to project', color='success', dismissable=True)
    except Exception as e:
        return dbc.Alert(f'Error: {e}', color='danger')

@app.callback(
    Output('pp-default-feedback','children'),
    Output('pp-table','data', allow_duplicate=True),
    Input('pp-make-default-btn','n_clicks'),
    State('pp-default-id','value'),
    prevent_initial_call=True
)
def make_default_profile(n, profile_id):
    if not n:
        raise PreventUpdate
    if not profile_id:
        return dbc.Alert('Profile ID required', color='danger'), dash.no_update
    try:
        # Set chosen default by marking is_default=1 and all others 0
        conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
        cur.execute('UPDATE pricing_profiles SET is_default=0')
        cur.execute('UPDATE pricing_profiles SET is_default=1 WHERE id=?',(int(profile_id),))
        conn.commit()
        cur.execute('SELECT id, name, sales_tax_rate, installation_rate, margin_multiplier, is_default FROM pricing_profiles ORDER BY id DESC')
        rows = [
            {
                'id':r[0], 'name':r[1],
                'sales_tax_rate': f"{(r[2] or 0)*100:.2f}%",
                'installation_rate': f"{(r[3] or 0)*100:.2f}%",
                'margin_multiplier': f"{r[4]:.3f}",
                'is_default': 'Yes' if r[5] else ''
            } for r in cur.fetchall()
        ]
        conn.close()
        return dbc.Alert('Default profile updated', color='success', dismissable=True), rows
    except Exception as e:
        return dbc.Alert(f'Error: {e}', color='danger'), dash.no_update

# Background autosync (database only) if enabled
if ONEDRIVE_SYNC_DIR and ONEDRIVE_AUTOSYNC_SEC > 0:
    import threading, time
    def _autosync_loop():
        while True:
            try:
                ok, msg = onedrive_manager.sync_database()
                if ok:
                    print(f"[autosync] {msg}")
            except Exception as e:
                print(f"[autosync][error] {e}")
            time.sleep(ONEDRIVE_AUTOSYNC_SEC)
    threading.Thread(target=_autosync_loop, daemon=True).start()

## Legacy single-image upload callback removed (replaced by multi-image system) – stray code block cleaned.

# Attempt secondary import from Book2.csv if dataset appears empty/minimal
def _attempt_import_book2():
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM sign_types")
        count = cur.fetchone()[0]
        conn.close()
        if count > 5:
            return
        csv_path = Path('Book2.csv')
        if not csv_path.exists():
            return
        print('[startup] Importing Book2.csv into sign_types...')
        import pandas as pd
        df = pd.read_csv(csv_path)
        # Normalize columns -> best effort mapping
        def parse_cost(val):
            if isinstance(val, str):
                return float(val.replace('$','').replace(',','') or 0) if any(ch.isdigit() for ch in val) else 0.0
            try: return float(val or 0)
            except: return 0.0
        records = []
        for r in df.to_dict('records'):
            name = str(r.get('Code') or r.get('Desc') or r.get('full_name') or '').strip()
            if not name:
                continue
            width = r.get('Width') or 0
            height = r.get('Height') or 0
            material = r.get('Material2') or r.get('Material') or ''
            # Derive price_per_sq_ft from 'Unnamed: 24' if numeric else material_multiplier
            ppsf_raw = r.get('Unnamed: 24') or r.get('material_multiplier') or 0
            # compute derived fields (example heuristic)
            try:
                ppsf = float(str(ppsf_raw).replace('$','').replace(',','')) if ppsf_raw not in (None,'') else 0.0
            except Exception:
                ppsf = 0.0
            records.append((name[:120], '', 0.0, material[:120], ppsf, float(width or 0), float(height or 0)))
        if records:
            conn = sqlite3.connect(DATABASE_PATH); cur = conn.cursor()
            cur.executemany('''INSERT OR IGNORE INTO sign_types (name, description, unit_price, material, price_per_sq_ft, width, height) VALUES (?,?,?,?,?,?,?)''', records)
            conn.commit(); conn.close()
            print(f"[startup] Imported {len(records)} records from Book2.csv")
    except Exception as e:
        print(f"[startup][import][warn] {e}")
# ---------------- Consolidated Role/Tab Globals (moved here after cleanup) ---------------- #
ROLE_OPTIONS = [
    {'label': 'Admin', 'value': 'admin'},
    {'label': 'Estimator', 'value': 'estimator'},
    {'label': 'Sales', 'value': 'sales'},
    {'label': 'Viewer', 'value': 'viewer'}
]
ROLE_RANK = {'admin':4,'estimator':3,'sales':2,'viewer':1}
TAB_DEFS = [
    ('projects-tab','Projects','viewer'),
    ('signs-tab','Sign Types','estimator'),
    ('groups-tab','Sign Groups','estimator'),
    ('building-tab','Building View','viewer'),
    ('estimates-tab','Estimates','viewer'),
    ('import-tab','Import Data','viewer'),
    ('tab_profiles','Pricing Profiles','admin'),
    ('tab_snapshots','Snapshots','estimator'),
    ('tab_templates','Bid Templates','estimator'),
    ('tab_tags','Tags','estimator'),
    ('tab_notes','Notes','estimator')
]
def build_tabs_for_role(role: str):
    rank = ROLE_RANK.get(role or 'viewer',1)
    tabs = []
    for tid,label,min_role in TAB_DEFS:
        if ROLE_RANK.get(min_role,1) <= rank:
            tabs.append(dbc.Tab(label=label, tab_id=tid))
    return tabs
# App layout
app.layout = dbc.Container([
    # Store for current role (session) and persisted last-used role (local)
    dcc.Store(id='current-role', storage_type='session'),
    dcc.Store(id='persisted-role', storage_type='local'),
    dcc.Store(id='persisted-active-tab', storage_type='local'),
    # Header with role selector
    dbc.Row([
        dbc.Col([
            html.Div([
                html.Img(src="/assets/LSI_Logo.svg", className="app-logo me-3"),
                html.H1("Sign Package Estimator", className="d-inline-block align-middle mb-0 me-4"),
                html.Div([
                    html.Span('Role:', className='me-2 fw-semibold'),
                    dcc.Dropdown(id='role-selector', options=ROLE_OPTIONS, value='viewer', clearable=False, style={'width':'180px','fontSize':'12px'})
                ], className='d-flex align-items-center')
            ], className="d-flex align-items-center py-3 flex-wrap")
        ])
    ]),
    dbc.Row([
    # Diagnostics banner removed per request
    # dbc.Col(html.Div(id='diagnostics-banner'))
    ]),
    # Dynamic tabs (role filtered)
    dbc.Row([
        dbc.Col([
            dbc.Tabs(id='main-tabs', active_tab='projects-tab')
        ])
    ], className='mb-3'),
    dcc.Interval(id='global-dropdown-refresh', interval=int(os.getenv('SIGN_APP_DROPDOWN_REFRESH_MS','300000')), n_intervals=0),
    dcc.Store(id='app-state', data={}),
    dcc.Store(id='last-error-message'),
    # dcc.Store for diagnostics banner dismissed (removed)
    # dcc.Store(id='env-banner-dismissed', data=False),
    dcc.Interval(id='status-refresh-interval', interval=5*60*1000, n_intervals=0),
    html.Div([
        dbc.Toast(id='app-error-toast', header='Notice', is_open=False, dismissable=True, duration=4000, icon='danger', style={'position':'fixed','top':10,'right':10,'zIndex':1080})
    ]),
    html.Div(id='tab-content', className='flex-grow-1'),
    html.Footer(
        className='app-footer text-center text-muted py-3 small mt-auto',
        children=[
            html.Span("© 2025 LSI Graphics, LLC"),
            html.Span(id='runtime-status', className='ms-3')
        ]
    )
], fluid=True, className='d-flex flex-column min-vh-100')

## (Removed temporary footer_status callback to avoid duplicate Output on runtime-status)

@app.callback(
    Output('role-selector','value'),            # ensure dropdown reflects restored/changed role
    Output('current-role','data'),              # session-scoped current role
    Output('main-tabs','children'),             # tabs for role
    Output('main-tabs','active_tab'),           # active tab (restored or first valid)
    Output('persisted-role','data'),            # update persisted role on change
    Input('role-selector','value'),             # user changing role OR default initial value
    State('persisted-role','data'),             # previously persisted role
    State('persisted-active-tab','data'),       # previously persisted active tab
    prevent_initial_call=False
)
def update_role_and_tabs(current_role_value, persisted_role, persisted_tab):
    """Unified bootstrap + role change handler.

    Logic:
    1. On first page load Dash will call with current_role_value = default ('viewer') and persisted states available.
       We prefer persisted role if present.
    2. Build tabs for the resolved role.
    3. Restore active tab if persisted and still valid for this role; else first tab id.
    4. When user manually changes role, persisted_role is used only for comparison; active tab resets to first unless previously persisted tab still valid.
    5. Persist the resolved role each invocation.
    """
    # Determine role to use (persisted overrides initial default if different and valid)
    resolved_role = current_role_value or 'viewer'
    if persisted_role and isinstance(persisted_role, dict):
        pr = persisted_role.get('role')
        if pr in {r['value'] for r in ROLE_OPTIONS} and pr != current_role_value and callback_context.triggered and 'role-selector' in callback_context.triggered[0]['prop_id']:
            # User actively changed role; keep selection not persisted override
            pass
        elif pr in {r['value'] for r in ROLE_OPTIONS} and (not callback_context.triggered or 'role-selector' not in callback_context.triggered[0]['prop_id']):
            # Initial load path: override with persisted
            resolved_role = pr
    tabs = build_tabs_for_role(resolved_role)
    tab_ids = [t.tab_id for t in tabs]
    # Determine desired active tab
    desired_tab = None
    if persisted_tab and isinstance(persisted_tab, dict):
        pt = persisted_tab.get('active_tab')
        if pt in tab_ids:
            desired_tab = pt
    # If user just changed the role (trigger from role-selector) and persisted tab invalid, fallback to first
    if not desired_tab:
        desired_tab = tab_ids[0] if tab_ids else None
    return resolved_role, resolved_role, tabs, desired_tab, {'role': resolved_role}

# Restore active tab on load once tabs are present
# We no longer need a separate restore_active_tab; init_role_and_tab handles both early.

# Persist active tab when user switches
@app.callback(
    Output('persisted-active-tab','data'),
    Input('main-tabs','active_tab'),
    prevent_initial_call=False
)
def persist_active_tab(active_tab):
    if not active_tab:
        raise PreventUpdate
    return {'active_tab': active_tab}

# If there is no persisted active tab yet when tabs first render, the above will save the default

def render_pricing_profiles_tab():
    return html.Div([
        html.H4('Pricing Profiles'),
        dbc.Row([
            dbc.Col([dbc.Label('Profile Name'), dbc.Input(id='pp-name', placeholder='Profile name')], md=3),
            dbc.Col([
                dbc.Label('Sales Tax Rate (%)'),
                dbc.Input(id='pp-tax', type='number', value=0),
                html.Small('Percent applied to taxable subtotal. Example: 7.5 = 7.5% tax.', className='text-muted d-block')
            ], md=2),
            dbc.Col([
                dbc.Label('Installation Rate (%)'),
                dbc.Input(id='pp-install', type='number', value=0),
                html.Small('Percent of subtotal added as install in percent mode.', className='text-muted d-block')
            ], md=2),
            dbc.Col([
                dbc.Label('Margin Multiplier'),
                dbc.Input(id='pp-margin', type='number', value=1.0, step=0.01),
                html.Small('Final total multiplier: subtotal * margin.', className='text-muted d-block')
            ], md=2),
            dbc.Col(dbc.Checklist(id='pp-default', options=[{'label':'Default','value':'d'}], value=[], className='mt-4'), md=1),
            dbc.Col(dbc.Button('Create Profile', id={'kind':'pp','action':'create'}, color='primary', className='mt-4 w-100'), md=2)
        ], className='g-2'),
        html.Div(id='pp-create-feedback', className='mt-2'),
        html.Hr(),
        dbc.Row([
            dbc.Col([
                html.H5('Profiles'),
                dash_table.DataTable(id='pp-table', columns=[
                    {'name':'ID','id':'id'}, {'name':'Name','id':'name'}, {'name':'Sales Tax','id':'sales_tax_rate'},
                    {'name':'Install Rate','id':'installation_rate'}, {'name':'Margin Mult','id':'margin_multiplier'}, {'name':'Default','id':'is_default'}
                ], data=[], page_size=8, style_table={'overflowX':'auto'})
            ], md=6),
            dbc.Col([
                html.H5('Assign to Project'),
                dbc.Row([
                    dbc.Col([dbc.Label('Project'), dcc.Dropdown(id='pp-project-id', placeholder='Select project')], md=4),
                    dbc.Col([dbc.Label('Profile'), dcc.Dropdown(id='pp-profile-id', placeholder='Select profile')], md=4),
                    dbc.Col(dbc.Button('Assign', id='pp-assign-btn', color='info', className='mt-4 w-100'), md=4)
                ], className='g-2'),
                html.Div(id='pp-assign-feedback', className='mt-2'),
                html.Hr(),
                html.H5('Set Default'),
                dbc.Row([
                    dbc.Col([dbc.Label('Profile'), dcc.Dropdown(id='pp-default-id', placeholder='Select profile')], md=5),
                    dbc.Col(dbc.Button('Make Default', id='pp-make-default-btn', color='secondary', className='mt-4 w-100'), md=5)
                ], className='g-2'),
                html.Div(id='pp-default-feedback', className='mt-2'),
                html.Hr(),
                dbc.Button('Refresh Profiles', id={'kind':'pp','action':'refresh'}, color='secondary', className='mt-2')
            ], md=6)
        ])
    ], style={'padding':'16px'})

def render_snapshots_tab():
    return html.Div([
        html.H4('Estimate Snapshots'),
        dbc.Row([
            dbc.Col([dbc.Label('Project'), dcc.Dropdown(id='snap-project-id', placeholder='Select project')], md=3),
            dbc.Col([dbc.Label('Label (optional)'), dbc.Input(id='snap-label', placeholder='Snapshot label')], md=3),
            dbc.Col(dbc.Button('Create Snapshot', id={'kind':'snapshot','action':'create'}, color='primary', className='mt-4 w-100'), md=2),
            dbc.Col(dbc.Button('Refresh List', id='refresh-snapshots-btn', color='secondary', className='mt-4 w-100'), md=2)
        ], className='g-2'),
        html.Div(id='snapshot-create-feedback', className='mt-2'),
        html.Hr(),
        dbc.Row([
            dbc.Col([
                html.H5('Snapshots'),
                dash_table.DataTable(id='snapshots-table', columns=[
                    {'name':'ID','id':'id'}, {'name':'Label','id':'label'}, {'name':'Hash','id':'snapshot_hash'}, {'name':'Created','id':'created_at'}
                ], data=[], page_size=8, style_table={'overflowX':'auto'})
            ], md=6),
            dbc.Col([
                html.H5('Diff'),
                dbc.Row([
                    dbc.Col([dbc.Label('Snapshot A'), dbc.Input(id='diff-snap-a', type='number')], md=3),
                    dbc.Col([dbc.Label('Snapshot B'), dbc.Input(id='diff-snap-b', type='number')], md=3),
                    dbc.Col(dbc.Button('Run Diff', id='run-diff-btn', color='info', className='mt-4 w-100'), md=2)
                ], className='g-2'),
                html.Pre(id='snapshot-diff-output', style={'maxHeight':'300px','overflowY':'auto','background':'#f8f9fa','padding':'8px','fontSize':'12px'})
            ], md=6)
        ])
    ], style={'padding':'16px'})

def render_templates_tab():
    return html.Div([
        html.H4('Bid Templates'),
        dbc.Row([
            dbc.Col([
                dbc.Label('Ordering'),
                dcc.RadioItems(id='template-order-mode', options=[{'label':'Newest','value':'new'},{'label':'A→Z','value':'alpha'}], value='new', inline=True, className='small')
            ], md=3)
        ], className='g-2 mb-1'),
        dbc.Row([
            dbc.Col([dbc.Label('Template Name'), dbc.Input(id='template-name', placeholder='Template name')], md=3),
            dbc.Col([dbc.Label('Description'), dbc.Input(id='template-desc', placeholder='Description')], md=4),
            dbc.Col(dbc.Button('Create Template', id={'kind':'template','action':'create'}, color='primary', className='mt-4 w-100'), md=2),
            dbc.Col(dbc.Button('Refresh', id={'kind':'template','action':'refresh'}, color='secondary', className='mt-4 w-100'), md=2)
        ], className='g-2'),
        html.Div(id='template-create-feedback', className='mt-2'),
        html.Hr(),
        dbc.Row([
            dbc.Col([
                html.H5('Templates'),
                dash_table.DataTable(id='templates-table', columns=[
                    {'name':'ID','id':'id'}, {'name':'Name','id':'name'}, {'name':'Description','id':'description'}, {'name':'Created','id':'created_at'}
                ], data=[], page_size=8, style_table={'overflowX':'auto'})
            ], md=5),
            dbc.Col([
                html.H5('Add Item to Template'),
                dbc.Row([
                    dbc.Col([dbc.Label('Template'), dcc.Dropdown(id='template-id-item', placeholder='Select template')], md=3),
                    dbc.Col([dbc.Label('Sign Type Name'), dbc.Input(id='template-sign-name', placeholder='Sign type name')], md=5),
                    dbc.Col([dbc.Label('Qty'), dbc.Input(id='template-sign-qty', type='number', value=1)], md=2),
                    dbc.Col(dbc.Button('Add Item', id='add-template-item-btn', color='success', className='mt-4 w-100'), md=2)
                ], className='g-2'),
                html.Div(id='template-item-feedback', className='mt-2'),
                html.Hr(),
                html.H5('Apply Template to Building'),
                dbc.Row([
                    dbc.Col([dbc.Label('Project'), dcc.Dropdown(id='apply-project-id', placeholder='Project')], md=3),
                    dbc.Col([dbc.Label('Template'), dcc.Dropdown(id='apply-template-id', placeholder='Template')], md=3),
                    dbc.Col([dbc.Label('Building'), dcc.Dropdown(id='apply-building-id', placeholder='Building', disabled=True)], md=3),
                    dbc.Col(dbc.Button('Apply', id='apply-template-btn', color='info', className='mt-4 w-100'), md=2)
                ], className='g-2'),
                html.Div(id='apply-template-feedback', className='mt-2')
            ], md=7)
        ])
    ], style={'padding':'16px'})

def render_tags_tab():
    return html.Div([
        html.H4('Sign Type Tags'),
        dbc.Row([
            dbc.Col([dbc.Label('Sign Type Name'), dbc.Input(id='tag-sign-type', placeholder='Sign type name')], md=4),
            dbc.Col([dbc.Label('Tag'), dbc.Input(id='tag-name', placeholder='Tag name')], md=3),
            dbc.Col(dbc.Button('Add Tag', id='add-tag-btn', color='primary', className='mt-4 w-100'), md=2),
            dbc.Col(dbc.Button('List Tags', id='list-tags-btn', color='secondary', className='mt-4 w-100'), md=2)
        ], className='g-2'),
        html.Div(id='tag-feedback', className='mt-2'),
        html.Pre(id='tag-list-output', style={'background':'#f8f9fa','padding':'8px','maxHeight':'260px','overflowY':'auto','fontSize':'12px'})
    ], style={'padding':'16px'})

def render_notes_tab():
    return html.Div([
        html.H4('Notes'),
        dbc.Row([
            dbc.Col([dbc.Label('Entity Type'), dcc.Dropdown(id='note-entity-type', options=[
                {'label':'Project','value':'project'}, {'label':'Building','value':'building'}, {'label':'Sign Type','value':'sign_type'}
            ], value='project', clearable=False, style={'width':'160px'})], md=2),
            dbc.Col([dbc.Label('Entity ID / Name'), dbc.Input(id='note-entity-id', placeholder='ID (project/building) or name (sign)')], md=3),
            dbc.Col([dbc.Label('Note Text'), dbc.Input(id='note-text', placeholder='Enter note')], md=4),
            dbc.Col(dbc.Checklist(id='note-include', options=[{'label':'Include in Export','value':'inc'}], value=['inc'], className='mt-4'), md=2),
            dbc.Col(dbc.Button('Add Note', id='add-note-btn', color='primary', className='mt-4 w-100'), md=1)
        ], className='g-2'),
        html.Div(id='note-add-feedback', className='mt-2'),
        html.Hr(),
        dbc.Row([
            dbc.Col([
                html.H5('Load Notes'),
                dbc.Row([
                    dbc.Col([dbc.Label('Entity Type'), dcc.Dropdown(id='list-note-entity-type', options=[
                        {'label':'Project','value':'project'}, {'label':'Building','value':'building'}, {'label':'Sign Type','value':'sign_type'}
                    ], value='project', clearable=False, style={'width':'160px'})], md=3),
                    dbc.Col([dbc.Label('Entity ID / Name'), dbc.Input(id='list-note-entity-id', placeholder='ID or name')], md=3),
                    dbc.Col(dbc.Button('Load Notes', id='load-notes-btn', color='secondary', className='mt-4 w-100'), md=2)
                ], className='g-2'),
                dash_table.DataTable(id='notes-table', columns=[
                    {'name':'ID','id':'id'}, {'name':'Note','id':'note'}, {'name':'Include','id':'include_in_export'}, {'name':'Created','id':'created_at'}
                ], data=[], page_size=10, style_table={'overflowX':'auto'})
            ], md=7),
            dbc.Col([
                html.H5('Toggle Include Flag'),
                dbc.Row([
                    dbc.Col([dbc.Label('Note ID'), dbc.Input(id='toggle-note-id', type='number')], md=5),
                    dbc.Col(dbc.Button('Toggle', id='toggle-note-btn', color='info', className='mt-4 w-100'), md=5)
                ], className='g-2'),
                html.Div(id='toggle-note-feedback', className='mt-2')
            ], md=5)
        ])
    ], style={'padding':'16px'})

@app.callback(
    Output('tab-content','children'),
    Input('main-tabs','active_tab'),
    State('current-role','data')
)
def render_active_tab(tab_id, role):
    try:
        if tab_id == 'projects-tab':
            return render_projects_tab()
        if tab_id == 'signs-tab':
            return render_signs_tab()
        if tab_id == 'groups-tab':
            return render_groups_tab()
        if tab_id == 'building-tab':
            return render_building_tab()
        if tab_id == 'estimates-tab':
            return render_estimates_tab()
        if tab_id == 'import-tab':
            return render_import_tab()
        if tab_id == 'tab_profiles':
            return render_pricing_profiles_tab()
        if tab_id == 'tab_snapshots':
            return render_snapshots_tab()
        if tab_id == 'tab_templates':
            return render_templates_tab()
        if tab_id == 'tab_tags':
            return render_tags_tab()
        if tab_id == 'tab_notes':
            return render_notes_tab()
        return html.Div('No tab content for selection.')
    except Exception as e:
        # Surface errors instead of leaving a blank area so issues are diagnosable when DEBUG=0
        return dbc.Alert([
            html.Strong('Tab render error: '),
            html.Code(str(e)[:260])
        ], color='danger', className='mt-3')

def get_project_tree_data():
    nodes = []
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        projects_df = pd.read_sql_query("SELECT id, name FROM projects", conn)
        for _, p in projects_df.iterrows():
            pid = f"project_{p['id']}"; nodes.append({'id':pid,'label':p['name'],'type':'project','level':0})
            buildings_df = pd.read_sql_query("SELECT id, name FROM buildings WHERE project_id=?", conn, params=(p['id'],))
            for _, b in buildings_df.iterrows():
                bid = f"building_{b['id']}"; nodes.append({'id':bid,'label':b['name'],'type':'building','level':1,'parent':pid})
                signs_df = pd.read_sql_query('''SELECT st.name, bs.quantity FROM building_signs bs JOIN sign_types st ON bs.sign_type_id=st.id WHERE bs.building_id=?''', conn, params=(b['id'],))
                for _, s in signs_df.iterrows():
                    nodes.append({'id':f"sign_{b['id']}_{s['name']}", 'label':f"{s['name']} ({s['quantity']})", 'type':'sign', 'level':2, 'parent':bid})
        conn.close()
    except Exception as e:
        print(f"[tree-data][warn] {e}")
    return nodes

def safe_tree_figure():
    try:
        nodes = get_project_tree_data()
        if not nodes:
            fig = go.Figure(); fig.update_layout(height=400, margin=dict(l=10,r=10,t=30,b=10)); return fig
        levels = {}
        for n in nodes: levels.setdefault(n['level'], []).append(n)
        x_gap = 240; y_gap = 42
        pos = {}
        for lvl, arr in levels.items():
            for i, n in enumerate(arr): pos[n['id']] = (lvl*x_gap, i*y_gap)
        edge_x=[]; edge_y=[]
        for n in nodes:
            if 'parent' in n and n['parent'] in pos:
                x0,y0 = pos[n['parent']]; x1,y1 = pos[n['id']]
                edge_x += [x0,x1,None]; edge_y += [y0,y1,None]
        fig = go.Figure()
        if edge_x:
            fig.add_trace(go.Scatter(x=edge_x,y=edge_y,mode='lines',line=dict(color='#cccccc',width=0.5),hoverinfo='none'))
        color_map={'project':'#1f77b4','building':'#ff7f0e','sign':'#2ca02c'}
        for lvl, arr in levels.items():
            fig.add_trace(go.Scatter(
                x=[pos[n['id']][0] for n in arr],
                y=[pos[n['id']][1] for n in arr],
                mode='markers+text',
                marker=dict(size=14,color=[color_map.get(n['type'],'#888') for n in arr]),
                text=[n['label'] for n in arr], textposition='middle right',
                hovertemplate='%{text}<extra></extra>', showlegend=False
            ))
        fig.update_layout(height=600, margin=dict(l=10,r=10,t=35,b=10), xaxis=dict(visible=False), yaxis=dict(visible=False))
        return fig
    except Exception as e:
        fig = go.Figure(); fig.add_annotation(text=f"Tree error: {e}", showarrow=False, x=0.5, y=0.5, xref='paper', yref='paper'); fig.update_layout(height=400)
        return fig

def render_projects_tab():
    """Render the projects management tab."""
    # Load current projects for initial render
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query("SELECT id, name, created_date FROM projects ORDER BY id DESC", conn)
    conn.close()
    if df.empty:
        project_list_component = html.Div("No projects yet.")
        project_options = []
    else:
        rows = [html.Li(f"{r.name} (ID {r.id}) - {r.created_date}") for r in df.itertuples()]
        project_list_component = html.Ul(rows, className="mb-0")
        project_options = [{"label": r.name, "value": r.id} for r in df.itertuples()]
    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H4("Project Tree Visualization")),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.RadioItems(
                                id='tree-view-mode',
                                options=[{'label':'Static','value':'static'},{'label':'Interactive','value':'cyto'}],
                                value='static', inline=True, className='mb-2'
                            )
                        ], md=8),
                        dbc.Col([
                            dbc.Button("Export Tree PNG", id='export-tree-png-btn', color='secondary', className='mt-1 w-100')
                        ], md=4)
                    ]),
                    # Interactive layout controls (shown only in Cytoscape mode)
                    html.Div(id='cyto-layout-controls', style={'display':'none'}, children=[
                        dbc.Row([
                            dbc.Col([
                                dbc.Label('Spacing Factor'),
                                dcc.Slider(id='cyto-spacing-factor', min=0.5, max=3.0, step=0.25, value=1.25,
                                           tooltip={'placement':'bottom','always_visible':False})
                            ], md=6),
                            dbc.Col([
                                dbc.Label('Padding'),
                                dcc.Slider(id='cyto-padding', min=0, max=100, step=5, value=15,
                                           tooltip={'placement':'bottom','always_visible':False})
                            ], md=6)
                        ], className='g-3 mb-2'),
                        dbc.Row([
                            dbc.Col([
                                dbc.Label('Label Max Width'),
                                dcc.Slider(id='cyto-label-width', min=60, max=220, step=10, value=120,
                                           tooltip={'placement':'bottom','always_visible':False})
                            ], md=6),
                            dbc.Col([
                                dbc.Label('Node Size'),
                                dcc.Slider(id='cyto-node-size', min=8, max=40, step=1, value=15,
                                           tooltip={'placement':'bottom','always_visible':False})
                            ], md=6)
                        ], className='g-3 mb-2'),
                        dbc.Row([
                            dbc.Col(dbc.Checklist(id='cyto-toggle-labels', options=[{'label':'Show Labels','value':'labels'}], value=['labels'], switch=True), md='auto'),
                            dbc.Col(dbc.Checklist(id='cyto-toggle-images', options=[{'label':'Show Images','value':'images'}], value=['images'], switch=True), md='auto'),
                            dbc.Col(dbc.Checklist(id='cyto-toggle-highlight', options=[{'label':'Highlight on Hover','value':'hl'}], value=['hl'], switch=True), md='auto')
                        ], className='g-3 mb-2 flex-wrap'),
                        html.Small('Adjust node spacing & layout padding for the interactive tree. Higher spacing factor spreads nodes further apart.', className='text-muted')
                    ]),
                    # Tree figure populated asynchronously by init callback; avoids layout-time exceptions blocking tab
                    html.Div(id='project-tree-wrapper', children=[
                        dcc.Graph(id='project-tree'),
                        # Hover panel for interactive (cyto) mode (initially hidden)
                        html.Div(id='cyto-hover-panel', style={'position':'absolute','top':'60px','right':'25px','zIndex':1050,'display':'none','maxWidth':'340px'})
                    ], style={'position':'relative'})
                ])
            ]),
            dcc.Download(id='tree-png-download'),
            dbc.Card([
                dbc.CardHeader(html.H4("Building & Sign Assignment")),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Select Project"),
                            dcc.Dropdown(id='assign-project-dropdown', placeholder='Choose project', options=project_options)
                        ], md=4),
                        dbc.Col([
                            dbc.Label("Create Building"),
                            dbc.Input(id='new-building-name', placeholder='Building name'),
                            dbc.Textarea(id='new-building-desc', placeholder='Description', style={"height": "60px"}),
                            dbc.Button("Add Building", id='add-building-btn', color='secondary', className='mt-2 w-100')
                        ], md=4),
                        dbc.Col([
                            dbc.Label("Select Building"),
                            dcc.Dropdown(id='building-dropdown', placeholder='Choose building'),
                            dbc.Input(id='rename-building-input', placeholder='Rename building', className='mt-2'),
                            dbc.Button('Rename', id='rename-building-btn', color='warning', size='sm', className='mt-1'),
                            html.Small(id='building-action-feedback', className='text-muted')
                        ], md=4)
                    ], className='mb-3 g-3'),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Add Sign Type"),
                            dcc.Dropdown(id='sign-type-dropdown', placeholder='Select sign type')
                        ], md=5),
                        dbc.Col([
                            dbc.Label("Quantity"),
                            dbc.Input(id='sign-qty-input', type='number', min=1, value=1)
                        ], md=3),
                        dbc.Col([
                            dbc.Button("Add / Update Sign", id='add-sign-to-building-btn', color='success', className='mt-4 w-100')
                        ], md=4)
                    ], className='mb-3'),
                    dash_table.DataTable(
                        id='building-signs-table',
                        columns=[
                            {"name": "Sign", "id": "sign_name", "editable": False},
                            {"name": "Quantity", "id": "quantity", "type": "numeric", "editable": True},
                            {"name": "Unit Price", "id": "unit_price", "type": "numeric", "editable": False},
                            {"name": "Total", "id": "total", "type": "numeric", "editable": False}
                        ],
                        data=[],
                        editable=True,
                        row_deletable=False,
                        style_table={"overflowX": "auto"},
                        page_size=10
                    ),
                    dbc.Button("Save Quantity Changes", id='save-building-signs-btn', color='primary', className='mt-2'),
                    html.Div(id='save-building-signs-feedback', className='mt-2'),
                    html.Hr(className='my-3'),
                    html.H5("Sign Groups for Building"),
                    dbc.Row([
                        dbc.Col(dbc.Checklist(id='project-groups-show-all', options=[{'label':'Show All Groups','value':'all'}], value=[], switch=True), width='auto')
                    ], className='mb-2 g-2'),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Add Group"),
                            dcc.Dropdown(id='project-group-dropdown', placeholder='Select sign group')
                        ], md=5),
                        dbc.Col([
                            dbc.Label("Quantity"),
                            dbc.Input(id='group-qty-input', type='number', min=1, value=1)
                        ], md=3),
                        dbc.Col([
                            dbc.Button("Add / Update Group", id='add-group-to-building-btn', color='success', className='mt-4 w-100')
                        ], md=4)
                    ], className='mb-3 g-3'),
                    dash_table.DataTable(
                        id='project-building-groups-table',
                        columns=[
                            {"name": "Group", "id": "group_name", "editable": False},
                            {"name": "Quantity", "id": "quantity", "type": "numeric", "editable": True}
                        ],
                        data=[],
                        editable=True,
                        row_deletable=False,
                        style_table={"overflowX":"auto"},
                        page_size=8
                    ),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Remove Group"),
                            dcc.Dropdown(id='assigned-group-delete-dropdown', placeholder='Assigned group')
                        ], md=8),
                        dbc.Col([
                            dbc.Button("Remove", id='remove-group-from-building-btn', color='danger', className='mt-4 w-100')
                        ], md=4)
                    ], className='mt-3 g-3'),
                    dbc.Button("Save Group Quantities", id='save-building-groups-btn', color='secondary', className='mt-2'),
                    html.Div(id='building-groups-feedback', className='mt-2')
                ])
            ], className='mt-3')
        ], width=8),
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H4("Project Management")),
                dbc.CardBody([
                    dbc.Form([
                        dbc.Row([
                            dbc.Label("Edit Existing", width=4),
                            dbc.Col(dcc.Dropdown(id='project-edit-dropdown', placeholder='Select project to edit', options=project_options), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Label("Debug Projects", width=4),
                            dbc.Col(html.Small(id='projects-debug-list', className='text-muted'), width=8)
                        ], className='mb-2'),
                        dbc.Row([
                            dbc.Label("Project Name", width=4),
                            dbc.Col(dbc.Input(id="project-name-input", type="text"), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Label("Description", width=4),
                            dbc.Col(dbc.Textarea(id="project-desc-input"), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Label("Sales Tax Rate (%)", width=4),
                            dbc.Col(dbc.Input(id="sales-tax-input", type="number", value=0), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Label("Installation Rate (%)", width=4),
                            dbc.Col(dbc.Input(id="installation-rate-input", type="number", value=0), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Label("Include Install?", width=4),
                            dbc.Col(dbc.Checklist(id="include-installation-input", options=[{"label": "Yes", "value": 1}], value=[1]), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Label("Include Sales Tax?", width=4),
                            dbc.Col(dbc.Checklist(id="include-sales-tax-input", options=[{"label": "Yes", "value": 1}], value=[1]), width=8)
                        ], className="mb-3"),
                        dbc.Row([
                            dbc.Col(dbc.Button("Create Project", id="create-project-btn", color="primary", className="w-100"), width=6),
                            dbc.Col(dbc.Button("Update Project", id="update-project-btn", color="secondary", className="w-100"), width=6)
                        ])
                    ], className='mb-2'),
                    dbc.Row([
                        dbc.Col(dbc.Button("Delete Project", id="delete-project-btn", color="danger", className="w-100"))
                    ]),
                    dcc.ConfirmDialog(id='delete-project-confirm', message='Delete this project and all related buildings, signs, and groups? This cannot be undone.')
                ])
            ], className="mb-3"),
            dbc.Card([
                dbc.CardHeader(html.H4("Existing Projects")),
                dbc.CardBody(id="projects-list", children=html.Div("No projects yet."))
            ]),
            html.Div(id="project-create-feedback")
        ], width=4)
    ])

def render_signs_tab():
    """Render the sign types management tab."""
    # Preload current sign_types so the table isn't empty if callback hasn't fired yet
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        preload_df = pd.read_sql_query(
            "SELECT name, description, material_alt, unit_price, material, price_per_sq_ft, material_multiplier, width, height, install_type, install_time_hours, per_sign_install_rate, image_path FROM sign_types ORDER BY name",
            conn
        )
        # Also preload material pricing so user immediately sees saved materials
        try:
            material_df = pd.read_sql_query("SELECT material_name, price_per_sq_ft FROM material_pricing ORDER BY material_name", conn)
        except Exception as me:
            print(f"[render_signs_tab][warn] failed preloading material_pricing: {me}")
            material_df = pd.DataFrame(columns=['material_name','price_per_sq_ft'])
        conn.close()
        preload_records = preload_df.to_dict('records')
        preload_materials = material_df.to_dict('records')
    except Exception as e:
        print(f"[render_signs_tab][warn] failed preloading sign_types: {e}")
        preload_records = []
        preload_materials = []
    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H4("Sign Types")),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Rows per page"),
                            dcc.Dropdown(
                                id='signs-page-size-dropdown',
                                options=[
                                    {'label':'10','value':10},
                                    {'label':'25','value':25},
                                    {'label':'50','value':50},
                                    {'label':'100','value':100},
                                    {'label':'All','value':-1}
                                ],
                                value=10,
                                clearable=False,
                                style={'width':'140px'}
                            )
                        ], width='auto'),
                        dbc.Col([
                            dbc.Label("Install Type"),
                            dcc.Dropdown(
                                id='signs-install-filter',
                                options=[
                                    {'label':'All','value':'all'},
                                    {'label':'Exterior (ext)','value':'ext'},
                                    {'label':'Non-Exterior','value':'non_ext'}
                                ],
                                value='all',
                                clearable=False,
                                style={'width':'170px'}
                            )
                        ], width='auto'),
                        dbc.Col(className='flex-grow-1')
                    ], className='g-3 mb-2 align-items-end'),
                    dash_table.DataTable(
                        id="signs-table",
                        columns=[
                            {"name": "Name", "id": "name", "editable": True},
                            {"name": "Description", "id": "description", "editable": True},
                            {"name": "Alt Material", "id": "material_alt", "editable": True},
                            {"name": "Material", "id": "material", "editable": True},
                            {"name": "Unit Price", "id": "unit_price", "type": "numeric", "editable": True},
                            {"name": "Price/Sq Ft", "id": "price_per_sq_ft", "type": "numeric", "editable": True},
                            {"name": "Mult", "id": "material_multiplier", "type": "numeric", "editable": True},
                            {"name": "Width", "id": "width", "type": "numeric", "editable": True},
                            {"name": "Height", "id": "height", "type": "numeric", "editable": True},
                            {"name": "Install Type", "id": "install_type", "editable": True},
                            {"name": "Install Hrs", "id": "install_time_hours", "type": "numeric", "editable": True},
                            {"name": "Per Sign Install $", "id": "per_sign_install_rate", "type": "numeric", "editable": True}
                        ],
                        data=preload_records,
                        editable=True,
                        row_deletable=True,
                        page_size=10,
                        style_table={"overflowX": "auto"}
                    ),
                    dbc.Button("Add New Sign Type", id="add-sign-btn", color="success", className="mt-3 me-2"),
                    dbc.Badge(id="signs-save-status", color="secondary", className="ms-2")
                ]),
                # Store original unfiltered sign types dataset to allow restoring after filters
                dcc.Store(id='signs-table-master-store', data=preload_records)
            ]),
            dbc.Card([
                dbc.CardHeader(html.H5("Sign Type Images")),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.Label('Select Sign Type'),
                            dcc.Dropdown(id='sign-image-sign-dropdown', options=[{'label':r['name'],'value':r['name']} for r in preload_records], placeholder='Choose sign type...')
                        ], width=6),
                        dbc.Col([
                            dbc.Label('Upload Images'),
                            dcc.Upload(
                                id='sign-image-upload',
                                children=html.Div(['Drag & Drop or ', html.A('Select Files')]),
                                multiple=True,
                                style={'border':'1px dashed #888','padding':'12px','textAlign':'center','cursor':'pointer'}
                            ),
                            html.Small('You can select multiple images. First upload sets cover if none present.', className='text-muted')
                        ], width=6)
                    ]),
                    html.Div(id='sign-image-upload-feedback', className='mt-2'),
                    html.Small('Supported: PNG/JPG/GIF/SVG. Stored under sign_images/. You can delete or set cover below.', className='text-muted'),
                    html.Hr(),
                    html.Div(id='sign-image-gallery')
                ])
            ], className='mt-3'),
            dbc.Card([
                dbc.CardHeader(html.H5("Material Pricing")),
                dbc.CardBody([
                    dash_table.DataTable(
                        id='material-pricing-table',
                        columns=[
                            {"name":"Material","id":"material_name","editable":True},
                            {"name":"Price / Sq Ft","id":"price_per_sq_ft","type":"numeric","editable":True}
                        ],
                        data=preload_materials, editable=True, row_deletable=True, page_size=8, style_table={'overflowX':'auto'}
                    ),
                    dbc.Button('Add Material', id='add-material-btn', color='secondary', className='mt-2 me-2'),
                    dbc.Button('Save Materials', id='save-materials-btn', color='primary', className='mt-2 me-2'),
                    dbc.Button('Recalculate Sign Prices', id='recalc-sign-prices-btn', color='warning', className='mt-2'),
                    html.Div(id='material-pricing-feedback', className='mt-2')
                ])
            ], className='mt-3')
        ]),
        dbc.Card([
            dbc.CardBody(
                dbc.Row([
                    dbc.Col(html.Small(id='debug-signs-stats', className='text-muted'), width='auto'),
                    dbc.Col(dbc.Button('×', id='close-debug-stats', color='link', size='sm', className='p-0 text-decoration-none'), width='auto')
                ], className='g-2 align-items-center justify-content-between flex-nowrap')
            )
        ], id='debug-signs-card', className='mt-2')
    ])

def render_groups_tab():
    """Render the sign groups management tab."""
    # Preload sign types & groups
    conn = sqlite3.connect(DATABASE_PATH)
    sign_types_df = pd.read_sql_query("SELECT id, name FROM sign_types ORDER BY name", conn)
    groups_df = pd.read_sql_query("SELECT id, name FROM sign_groups ORDER BY name", conn)
    conn.close()
    sign_type_options = [{"label": r.name, "value": r.id} for r in sign_types_df.itertuples()]
    group_options = [{"label": r.name, "value": r.id} for r in groups_df.itertuples()]
    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H5("Create / Edit Group")),
                dbc.CardBody([
                    dbc.Input(id='group-name-input', placeholder='Group name'),
                    dbc.Textarea(id='group-desc-input', placeholder='Description', className='mt-2', style={'height': '60px'}),
                    dbc.Button('Save Group', id='save-group-btn', color='primary', className='mt-2 w-100'),
                    html.Div(id='group-save-feedback', className='mt-2')
                ])
            ]),
            dbc.Card([
                dbc.CardHeader(html.H5("Existing Groups")),
                dbc.CardBody([
                    dcc.Dropdown(id='group-select-dropdown', options=group_options, placeholder='Select group'),
                    dash_table.DataTable(
                        id='group-members-table',
                        columns=[
                            {"name": "Sign", "id": "sign_name"},
                            {"name": "Quantity", "id": "quantity", "type": "numeric"}
                        ],
                        data=[], editable=True, row_deletable=True, page_size=8, style_table={'overflowX': 'auto'}
                    ),
                    dbc.Row([
                        dbc.Col(dcc.Dropdown(id='group-add-sign-dropdown', options=sign_type_options, placeholder='Add sign type'), width=7),
                        dbc.Col(dbc.Input(id='group-add-sign-qty', type='number', min=1, value=1), width=3),
                        dbc.Col(dbc.Button('Add', id='group-add-sign-btn', color='success', className='w-100'), width=2)
                    ], className='mt-2 g-2'),
                    dbc.Button('Save Member Changes', id='group-save-members-btn', color='secondary', className='mt-2'),
                    html.Div(id='group-members-feedback', className='mt-2')
                ])
            ], className='mt-3')
        ], width=6),
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H5("Assign Groups to Building")),
                dbc.CardBody([
                    dcc.Dropdown(id='group-assign-project-dropdown', placeholder='Select project'),
                    dcc.Dropdown(id='group-assign-building-dropdown', placeholder='Select building', className='mt-2'),
                    dcc.Dropdown(id='group-assign-group-dropdown', options=group_options, placeholder='Select group', className='mt-2'),
                    dbc.Input(id='group-assign-qty', type='number', min=1, value=1, className='mt-2'),
                    dbc.Button('Add Group to Building', id='group-assign-btn', color='success', className='mt-2 w-100'),
                    html.Div(id='group-assign-feedback', className='mt-2'),
                    dash_table.DataTable(
                        id='building-groups-table',
                        columns=[
                            {"name": "Group", "id": "group_name"},
                            {"name": "Quantity", "id": "quantity", "type": "numeric"}
                        ],
                        data=[], editable=True, row_deletable=False, page_size=8
                    ),
                    dbc.Button('Save Group Quantities', id='building-save-group-qty-btn', color='primary', className='mt-2'),
                    html.Div(id='building-group-save-feedback', className='mt-2')
                ])
            ])
        ], width=6)
    ])

def render_building_tab():
    """Render dedicated building view and sign management."""
    # Preload project options
    conn = sqlite3.connect(DATABASE_PATH)
    pdf = pd.read_sql_query('SELECT id, name FROM projects ORDER BY name', conn)
    conn.close()
    project_options = ([{'label': r.name, 'value': r.id} for r in pdf.itertuples()]) if not pdf.empty else []
    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H4('Select Building')),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.Label('Project'),
                            dcc.Dropdown(id='bv-project-dropdown', options=project_options, placeholder='Select project')
                        ], md=4),
                        dbc.Col([
                            dbc.Label('Building'),
                            dcc.Dropdown(id='bv-building-dropdown', placeholder='Select building')
                        ], md=4),
                        dbc.Col([
                            dbc.Label('Rename'),
                            dbc.Input(id='bv-rename-building-input', placeholder='New building name'),
                            dbc.Button('Apply', id='bv-rename-building-btn', size='sm', color='warning', className='mt-1 w-100')
                        ], md=4)
                    ], className='g-3'),
                    html.Div(id='bv-building-meta', className='text-muted small mt-2')
                ])
            ]),
            dbc.Card([
                dbc.CardHeader(html.H5('Manage Signs')),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.Label('Add / Update Sign Type'),
                            dcc.Dropdown(id='bv-sign-type-dropdown', placeholder='Select sign type')
                        ], md=5),
                        dbc.Col([
                            dbc.Label('Quantity'),
                            dbc.Input(id='bv-sign-qty-input', type='number', min=1, value=1)
                        ], md=3),
                        dbc.Col([
                            dbc.Button('Add / Update', id='bv-add-sign-btn', color='success', className='mt-4 w-100')
                        ], md=4)
                    ], className='g-3 mb-3'),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label('Delete Sign'),
                            dcc.Dropdown(id='bv-delete-sign-dropdown', placeholder='Assigned sign')
                        ], md=8),
                        dbc.Col([
                            dbc.Button('Delete', id='bv-delete-sign-btn', color='danger', className='mt-4 w-100')
                        ], md=4)
                    ], className='g-3 mb-2'),
                    html.Hr(),
                    dbc.Row([
                        dbc.Col(html.H6('Groups', className='mt-2'), width=12)
                    ], className='mb-1'),
                    dbc.Row([
                        dbc.Col([
                            dbc.Label('Remove Group'),
                            dcc.Dropdown(id='bv-delete-group-dropdown', placeholder='Assigned group')
                        ], md=8),
                        dbc.Col([
                            dbc.Button('Remove Group', id='bv-delete-group-btn', color='danger', className='mt-4 w-100')
                        ], md=4)
                    ], className='g-3 mb-2'),
                    dash_table.DataTable(
                        id='bv-signs-table',
                        columns=[
                            {'name':'Sign','id':'sign_name','editable':False},
                            {'name':'Quantity','id':'quantity','type':'numeric','editable':True},
                            {'name':'Unit Price','id':'unit_price','type':'numeric','editable':False},
                            {'name':'Total','id':'total','type':'numeric','editable':False}
                        ],
                        data=[], editable=True, row_deletable=False, page_size=12, style_table={'overflowX':'auto'}
                    ),
                    dbc.Button('Save Quantities', id='bv-save-signs-btn', color='primary', className='mt-2'),
                    html.Div(id='bv-feedback', className='mt-2')
                ])
            ], className='mt-3')
        ], width=8),
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H5('Summary')),
                dbc.CardBody([
                    html.Div(id='bv-summary', className='fw-bold'),
                    html.Hr(),
                    html.Small('Modify sign assignments directly. Use Projects tab for groups.', className='text-muted')
                ])
            ])
        ], width=4)
    ])

def render_estimates_tab():
    """Render the cost estimation and export tab."""
    # Populate project options fresh on each render
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query("SELECT id, name FROM projects ORDER BY name", conn)
    bdf = pd.read_sql_query("SELECT b.id, b.name, b.project_id, p.name as project_name FROM buildings b JOIN projects p ON b.project_id=p.id ORDER BY p.name, b.name", conn)
    conn.close()
    project_options = [{"label": r.name, "value": r.id} for r in df.itertuples()] if not df.empty else []
    building_options = [{"label": f"{r.name} (Project: {r.project_name})", "value": r.id} for r in bdf.itertuples()] if not bdf.empty else []
    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H4("Generate Project Estimate")),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            dbc.Label("Project"),
                            dcc.Dropdown(id='estimate-project-dropdown', placeholder='Select project', options=project_options)
                        ], md=4),
                        dbc.Col([
                            dbc.Label("Or Single Building"),
                            dcc.Dropdown(id='estimate-building-dropdown', placeholder='Optional: choose building(s)', options=building_options, multi=True)
                        ], md=4),
                        dbc.Col([
                            dbc.Button("Generate Estimate", id='generate-estimate-btn', color='primary', className='mt-4 w-100')
                        ], md=2),
                        dbc.Col([
                            dbc.Button("Export to Excel", id='export-estimate-btn', color='success', className='mt-4 w-100', disabled=True)
                        ], md=2),
                        dbc.Col([
                            dbc.Button("Export PDF", id='export-estimate-pdf-btn', color='secondary', className='mt-4 w-100', disabled=True)
                        ], md=2),
                        dbc.Col([
                            dbc.Input(id='estimate-pdf-title', placeholder='PDF Title (optional)', className='mt-4')
                        ], md=3),
                        dbc.Col([
                            dcc.Download(id='estimate-download'),
                            dcc.Download(id='estimate-pdf-download')
                        ], md=2),
                        dbc.Col([
                            dbc.Checklist(
                                id='embed-images-toggle',
                                options=[{'label':'Embed Images in Exports','value':'embed'}],
                                value=['embed'],
                                className='mt-4',
                                switch=True
                            )
                        ], md=3),
                        dbc.Col([
                            dbc.Checklist(
                                id='exterior-only-toggle',
                                options=[{'label':'Exterior Only','value':'ext_only'}],
                                value=[],
                                className='mt-4',
                                switch=True
                            )
                        ], md=2),
                        dbc.Col([
                            dbc.Checklist(
                                id='non-exterior-only-toggle',
                                options=[{'label':'Non-Exterior Only','value':'non_ext_only'}],
                                value=[],
                                className='mt-4',
                                switch=True
                            )
                        ], md=2),
                        dbc.Col([
                            dbc.Checklist(
                                id='disable-logo-toggle',
                                options=[{'label':'No Logo','value':'no_logo'}],
                                value=[],
                                className='mt-4',
                                switch=True
                            )
                        ], md=2)
                    ], className='mb-3 g-2'),
                    dcc.Store(id='embed-images-store', data={'embed': True}),
                    dcc.Store(id='disable-logo-store', data={'disable': False}),
                    dbc.Accordion([
                        dbc.AccordionItem([
                            dbc.Row([
                                dbc.Col([
                                    dbc.Label('Sign Price Basis'),
                                    dbc.RadioItems(id='price-calc-mode', options=[
                                        {'label':'Stored Unit Price (Per Sign)','value':'per_sign'},
                                        {'label':'Area * Material Rate','value':'per_area'}
                                    ], value='per_sign', inline=False)
                                ], md=4),
                                dbc.Col([
                                    dbc.Label('Installation Mode'),
                                    dbc.RadioItems(id='install-mode', options=[
                                        {'label':'Percent of Subtotal','value':'percent'},
                                        {'label':'Per Sign','value':'per_sign'},
                                        {'label':'Per Area (sq ft)','value':'per_area'},
                                        {'label':'Hours * Hourly Rate','value':'hours'},
                                        {'label':'None','value':'none'}
                                    ], value='percent', inline=False)
                                ], md=4),
                                dbc.Col([
                                    dbc.Label('Parameters'),
                                    dbc.Row([
                                        dbc.Col(dbc.Input(id='install-percent-input', type='number', value=0, placeholder='% Install'), width=6),
                                        dbc.Col(dbc.Input(id='install-per-sign-rate', type='number', value=0, placeholder='Install $/Sign'), width=6)
                                    ], className='g-1 mt-1'),
                                    dbc.Row([
                                        dbc.Col(dbc.Input(id='install-per-area-rate', type='number', value=0, placeholder='Install $/SqFt'), width=6),
                                        dbc.Col(dbc.Input(id='install-hours-input', type='number', value=0, placeholder='Hours'), width=3),
                                        dbc.Col(dbc.Input(id='install-hourly-rate-input', type='number', value=0, placeholder='$/Hour'), width=3)
                                    ], className='g-1 mt-1'),
                                    dbc.Checklist(id='auto-install-use', options=[{'label':'Use auto per-sign rates / hours when manual blank','value':1}], value=[1], className='mt-2'),
                                    html.Small("Only the fields relevant to selected modes are used.", className='text-muted')
                                ], md=4)
                            ], className='g-3')
                        ], title='Advanced Calculation Options')
                    ], start_collapsed=True, className='mb-3'),
                    dash_table.DataTable(
                        id='estimate-table',
                        columns=[
                            {"name": "Building", "id": "Building"},
                            {"name": "Item", "id": "Item"},
                            {"name": "Material", "id": "Material"},
                            {"name": "Dimensions", "id": "Dimensions"},
                            {"name": "Quantity", "id": "Quantity", "type": "numeric"},
                            {"name": "Unit Price", "id": "Unit_Price", "type": "numeric"},
                            {"name": "Total", "id": "Total", "type": "numeric"}
                        ],
                        data=[],
                        style_table={"overflowX": "auto"},
                        page_size=15
                    ),
                    html.Hr(),
                    html.Div(id='estimate-summary')
                ])
            ])
        ])
    ])

def render_import_tab():
    """Render the data import tab."""
    return dbc.Row([
        dbc.Col([
            dbc.Card([
                dbc.CardHeader(html.H4("Import CSV Data")),
                dbc.CardBody([
                    dcc.Upload(
                        id='upload-data',
                        children=html.Div([
                            'Drag and Drop or ',
                            html.A('Select Files')
                        ]),
                        style={
                            'width': '100%',
                            'height': '60px',
                            'lineHeight': '60px',
                            'borderWidth': '1px',
                            'borderStyle': 'dashed',
                            'borderRadius': '5px',
                            'textAlign': 'center',
                            'margin': '10px'
                        },
                        multiple=False
                    ),
                    html.Div(id='upload-output'),
                    html.Hr(),
                    dbc.Button('Import Local Book2.csv', id='import-book2-btn', color='secondary'),
                    html.Div(id='book2-import-feedback', className='mt-2 text-muted', style={'fontSize':'0.9rem'})
                ])
            ])
        ])
    ])

# ------------------ Embed Images Toggle Sync ------------------ #
@app.callback(
    Output('embed-images-store','data'),
    Input('embed-images-toggle','value'),
    prevent_initial_call=False
)
def sync_embed_toggle(values):
    # values is list like ['embed'] when on, or [] when off
    try:
        embed_on = 'embed' in (values or [])
    except Exception:
        embed_on = True
    return {'embed': embed_on}

# ------------------ Sign Type Multi-Image Management ------------------ #
@app.callback(
    Output('sign-image-upload-feedback','children'),
    Output('sign-image-gallery','children'),
    Input('sign-image-upload','contents'),
    State('sign-image-upload','filename'),
    State('sign-image-sign-dropdown','value'),
    prevent_initial_call=True
)
def handle_sign_type_multi_image(contents_list, filenames, sign_name):
    if not contents_list or not sign_name:
        raise PreventUpdate
    if not isinstance(contents_list, list):
        # Dash may pass single as string
        contents_list = [contents_list]
        filenames = [filenames]
    saved = 0
    errors = []
    from hashlib import sha1
    img_dir = Path('sign_images'); img_dir.mkdir(exist_ok=True)
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    # fetch sign_type id
    cur.execute('SELECT id, image_path FROM sign_types WHERE name=?', (sign_name,))
    row = cur.fetchone()
    if not row:
        conn.close(); return dbc.Alert('Sign type not found', color='danger'), dash.no_update
    sign_type_id, cover_path = row
    try:
        cur.execute('SELECT COUNT(*) FROM sign_type_images WHERE sign_type_id=?', (sign_type_id,))
        existing_count = cur.fetchone()[0]
    except Exception:
        existing_count = 0
    for content, fname in zip(contents_list, filenames):
        try:
            header, b64data = content.split(',',1)
            raw = base64.b64decode(b64data)
            ext = ''.join(Path(fname).suffix.split()) or '.bin'
            digest = sha1(raw).hexdigest()[:16]
            out_name = f"{sign_type_id}_{digest}{ext}"
            out_path = img_dir / out_name
            out_path.write_bytes(raw)
            # insert into sign_type_images
            cur.execute('INSERT INTO sign_type_images (sign_type_id, image_path, display_order, created_at, file_hash) VALUES (?,?,?,?,?)',
                        (sign_type_id, str(out_path), existing_count + saved + 1, datetime.now().isoformat(timespec='seconds'), digest))
            # set cover if none or first image and sign_types.image_path empty
            if not cover_path:
                cur.execute('UPDATE sign_types SET image_path=? WHERE id=?', (str(out_path), sign_type_id))
                cover_path = str(out_path)
            saved += 1
        except Exception as e:
            errors.append(f"{fname}: {e}")
    conn.commit(); conn.close()
    feedback_children = []
    if saved:
        feedback_children.append(dbc.Alert(f"Uploaded {saved} image(s) for {sign_name}", color='success'))
    if errors:
        feedback_children.append(dbc.Alert('Errors: ' + '; '.join(errors), color='warning'))
    return feedback_children, _render_sign_image_gallery(sign_name)


@app.callback(
    Output('sign-image-gallery','children', allow_duplicate=True),
    Input('sign-image-sign-dropdown','value'),
    prevent_initial_call=True
)
def refresh_gallery_on_select(sign_name):
    if not sign_name:
        raise PreventUpdate
    return _render_sign_image_gallery(sign_name)


@app.callback(
    Output('sign-image-gallery','children', allow_duplicate=True),
    Input({'action':'sign-image-set-cover','path':dash.ALL}, 'n_clicks'),
    State('sign-image-sign-dropdown','value'),
    prevent_initial_call=True
)
def set_cover_image(n_clicks_list, sign_name):
    ctx = dash.callback_context
    if not ctx.triggered or not sign_name:
        raise PreventUpdate
    trig = ctx.triggered[0]['prop_id'].split('.')[0]
    try:
        import json
        meta = json.loads(trig)
        path = meta.get('path')
    except Exception:
        raise PreventUpdate
    if not path:
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    cur.execute('SELECT id FROM sign_types WHERE name=?', (sign_name,))
    row = cur.fetchone()
    if not row:
        conn.close(); raise PreventUpdate
    sign_type_id = row[0]
    cur.execute('UPDATE sign_types SET image_path=? WHERE id=?', (path, sign_type_id))
    conn.commit(); conn.close()
    return _render_sign_image_gallery(sign_name)


def _render_sign_image_gallery(sign_name: str):
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    cur.execute('SELECT id, image_path, display_order FROM sign_type_images sti JOIN sign_types st ON sti.sign_type_id=st.id WHERE st.name=? ORDER BY display_order', (sign_name,))
    rows = cur.fetchall()
    cur.execute('SELECT image_path FROM sign_types WHERE name=?', (sign_name,))
    cover = (cur.fetchone() or [None])[0]
    conn.close()
    if not rows:
        return html.Div(html.I('No images uploaded yet.'), className='text-muted')
    cards = []
    for _id, path, order in rows:
        is_cover = (str(path) == str(cover))
        actions = [
            dbc.Button('Set Cover', id={'action':'sign-image-set-cover','path':path}, color='secondary', size='sm', className='me-1', disabled=is_cover),
            # Delete button handled in future callback (todo #3)
            dbc.Button('Delete', id={'action':'sign-image-delete','path':path}, color='danger', size='sm', outline=True)
        ]
        style_border = '2px solid #157347' if is_cover else '1px solid #ccc'
        cards.append(
            dbc.Col(dbc.Card([
                html.Div(html.Img(src=convert_path_to_data_url(path), style={'maxWidth':'100%','maxHeight':'120px','objectFit':'contain'}), className='p-2'),
                dbc.CardFooter([
                    html.Small(Path(path).name, className='d-block text-truncate', style={'maxWidth':'140px'}),
                    html.Div(actions, className='mt-1')
                ], style={'background':'#fafafa'})
            ], style={'border':style_border,'boxShadow':'0 1px 2px rgba(0,0,0,0.08)'}, className='h-100'), width=3)
        )
    return dbc.Row(cards, className='g-2')

def convert_path_to_data_url(path: str):
    try:
        p = Path(path)
        if not p.exists():
            return ''
        data = p.read_bytes()
        import mimetypes, base64
        mt, _ = mimetypes.guess_type(p.name)
        if not mt:
            mt = 'application/octet-stream'
        return f"data:{mt};base64,{base64.b64encode(data).decode()}"
    except Exception:
        return ''

# Delete image callback (removes DB row + file and reassigns cover if needed)
@app.callback(
    Output('sign-image-gallery','children', allow_duplicate=True),
    Input({'action':'sign-image-delete','path':dash.ALL}, 'n_clicks'),
    State('sign-image-sign-dropdown','value'),
    prevent_initial_call=True
)
def delete_sign_type_image(n_clicks_list, sign_name):
    ctx = dash.callback_context
    if not ctx.triggered or not sign_name:
        raise PreventUpdate
    trig = ctx.triggered[0]['prop_id'].split('.')[0]
    try:
        import json
        meta = json.loads(trig)
        path = meta.get('path')
    except Exception:
        raise PreventUpdate
    if not path:
        raise PreventUpdate
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        # Get sign type id & current cover
        cur.execute('SELECT id, image_path FROM sign_types WHERE name=?', (sign_name,))
        row = cur.fetchone()
        if not row:
            conn.close(); raise PreventUpdate
        sign_type_id, cover_path = row
        # Delete the image row
        cur.execute('DELETE FROM sign_type_images WHERE sign_type_id=? AND image_path=?', (sign_type_id, path))
        # If it was cover, choose the next lowest display_order
        if cover_path and str(cover_path) == str(path):
            cur.execute('SELECT image_path FROM sign_type_images WHERE sign_type_id=? ORDER BY display_order LIMIT 1', (sign_type_id,))
            next_cover = cur.fetchone()
            new_cover = next_cover[0] if next_cover else None
            cur.execute('UPDATE sign_types SET image_path=? WHERE id=?', (new_cover, sign_type_id))
        conn.commit(); conn.close()
    except Exception as e:
        print(f"[sign-image-delete][error] {e}")
    # Delete file from filesystem (best-effort)
    try:
        fp = Path(path)
        if fp.exists():
            fp.unlink()
    except Exception as fe:
        print(f"[sign-image-delete][file][warn] {fe}")
    return _render_sign_image_gallery(sign_name)

@app.callback(
    Output('upload-output', 'children'),
    Output('signs-table', 'data', allow_duplicate=True),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename'),
    prevent_initial_call=True
)
def update_output(contents, filename):
    if contents is not None:
        # Process uploaded file
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)
        
        try:
            # Assume CSV file
            df = pd.read_csv(io.StringIO(decoded.decode('utf-8')))
            # Minimal inline loader (previous helper removed during refactor)
            conn = sqlite3.connect(DATABASE_PATH)
            cur = conn.cursor()
            inserted = 0
            for r in df.to_dict('records'):
                name = str(r.get('name') or r.get('Code') or r.get('Desc') or '').strip()
                if not name:
                    continue
                unit_price = r.get('unit_price') or r.get('price') or 0
                material = r.get('material') or r.get('Material') or ''
                try:
                    cur.execute('''
                        INSERT INTO sign_types (name, description, unit_price, material)
                        VALUES (?,?,?,?)
                        ON CONFLICT(name) DO UPDATE SET description=excluded.description, unit_price=excluded.unit_price, material=excluded.material
                    ''', (
                        name[:120], str(r.get('description') or r.get('Desc') or '')[:255], float(unit_price or 0), str(material)[:120]
                    ))
                    inserted += 1
                except Exception as ie:
                    print(f"[import][warn] row skipped: {ie}")
            conn.commit(); conn.close()
            # Reload table data after import
            conn = sqlite3.connect(DATABASE_PATH)
            table_df = pd.read_sql_query("SELECT name, description, unit_price, material, price_per_sq_ft, width, height FROM sign_types ORDER BY name", conn)
            conn.close()
            return dbc.Alert(f"Imported/updated {inserted} sign types from {filename}", color="success"), table_df.to_dict('records')
            
        except Exception as e:
            return dbc.Alert(f"Error processing file: {str(e)}", color="danger"), dash.no_update
    
    return html.Div(), dash.no_update

# ----------- Force Import Book2.csv (manual trigger) ------------- #
def _force_import_book2(csv_path: Path):
    if not csv_path.exists():
        raise FileNotFoundError(f"{csv_path} not found")
    df = pd.read_csv(csv_path)
    records = []
    def parse_cost(val):
        if isinstance(val, str):
            txt = val.replace('$','').replace(',','').strip()
            if not txt:
                return 0.0
            try:
                return float(txt)
            except:
                return 0.0
        try:
            return float(val or 0)
        except:
            return 0.0
    for r in df.to_dict('records'):
        name = str(r.get('Code') or r.get('Desc') or r.get('full_name') or '').strip()
        if not name:
            continue
        width = r.get('Width') or 0
        height = r.get('Height') or 0
        material = r.get('Material2') or r.get('Material') or ''
        ppsf_raw = r.get('material_multiplier') or r.get('Unnamed: 24') or 0
        try:
            ppsf = float(str(ppsf_raw).replace('$','').replace(',','')) if ppsf_raw not in (None,'') else 0.0
        except:
            ppsf = 0.0
        unit_price = parse_cost(r.get('item_cost'))
        try:
            if unit_price == 0 and width and height and ppsf:
                unit_price = float(width) * float(height) * float(ppsf)
        except:
            pass
        records.append((name[:120], str(r.get('Desc') or r.get('full_name') or '')[:255], unit_price, str(material)[:120], ppsf, float(width or 0), float(height or 0)))
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    cur.executemany('''
        INSERT INTO sign_types (name, description, unit_price, material, price_per_sq_ft, width, height)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(name) DO UPDATE SET description=excluded.description,
            unit_price=excluded.unit_price, material=excluded.material,
            price_per_sq_ft=excluded.price_per_sq_ft, width=excluded.width, height=excluded.height
    ''', records)
    conn.commit()
    cur.execute('SELECT COUNT(*) FROM sign_types')
    total = cur.fetchone()[0]
    conn.close()
    return len(records), total

@app.callback(
    Output('book2-import-feedback','children'),
    Output('signs-table','data', allow_duplicate=True),
    Input('import-book2-btn','n_clicks'),
    prevent_initial_call=True
)
def manual_import_book2(n):
    if not n:
        raise PreventUpdate
    csv_path = Path('Book2.csv')
    try:
        imported, total = _force_import_book2(csv_path)
        conn = sqlite3.connect(DATABASE_PATH)
        df = pd.read_sql_query(
            "SELECT name, description, material_alt, unit_price, material, price_per_sq_ft, material_multiplier, width, height, install_type, install_time_hours, per_sign_install_rate, image_path FROM sign_types ORDER BY name",
            conn
        )
        conn.close()
        return dbc.Alert(f"Imported/merged {imported} rows from {csv_path.name}. sign_types now has {total} rows.", color='success'), df.to_dict('records')
    except Exception as e:
        return dbc.Alert(f"Import failed: {e}", color='danger'), dash.no_update

# ------------------ Project Creation & Listing ------------------ #
@app.callback(
    Output('projects-list', 'children', allow_duplicate=True),
    Output('project-create-feedback', 'children', allow_duplicate=True),
    Output('project-tree', 'figure', allow_duplicate=True),
    Output('assign-project-dropdown', 'options'),
    Output('project-edit-dropdown', 'options', allow_duplicate=True),
    Output('projects-debug-list','children', allow_duplicate=True),
    Input('create-project-btn', 'n_clicks'),
    Input('main-tabs','active_tab'),
    State('project-name-input', 'value'),
    State('project-desc-input', 'value'),
    State('sales-tax-input', 'value'),
    State('installation-rate-input', 'value'),
    State('include-installation-input', 'value'),
    State('include-sales-tax-input', 'value'),
    prevent_initial_call=True
)
def create_or_refresh_projects(n_clicks, active_tab, name, desc, sales_tax, install_rate, include_install_values, include_tax_values):
    """Create a project or hydrate existing list when Projects tab first shown.

    Logic:
    - If triggered by tab activation (active_tab == 'projects-tab') and no click happened yet, just load existing projects.
    - If triggered by button click, attempt to create then reload list.
    """
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    hydrate_only = ('main-tabs' in triggered and active_tab == 'projects-tab' and not n_clicks)
    create_mode = ('create-project-btn' in triggered)
    if not (hydrate_only or create_mode):
        raise PreventUpdate
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        feedback = dash.no_update
        if create_mode:
            if not name:
                return (dash.no_update, dbc.Alert("Project name required", color='danger'), dash.no_update, dash.no_update, dash.no_update, dash.no_update)
            try:
                cur.execute("INSERT INTO projects (name, description, sales_tax_rate, installation_rate, include_installation, include_sales_tax) VALUES (?,?,?,?,?,?)", (
                    name.strip(),
                    desc or '',
                    float(sales_tax or 0)/100.0,
                    float(install_rate or 0)/100.0,
                    1 if (include_install_values and 1 in include_install_values) else 0,
                    1 if (include_tax_values and 1 in include_tax_values) else 0
                ))
                conn.commit()
                feedback = dbc.Alert(f"Project '{name}' created", color='success', dismissable=True)
            except sqlite3.IntegrityError:
                feedback = dbc.Alert(f"Project '{name}' already exists", color='warning')
            except Exception as e:
                return dash.no_update, dbc.Alert(f"Error: {e}", color='danger'), dash.no_update, dash.no_update, dash.no_update, dash.no_update
        df = pd.read_sql_query("SELECT id, name, created_date FROM projects ORDER BY id DESC", conn)
        conn.close()
        if df.empty:
            list_children = html.Div("No projects yet.")
            project_options = []
            debug_txt = 'none'
        else:
            rows = [html.Li(f"{r.name} (ID {r.id}) - {r.created_date}") for r in df.itertuples()]
            list_children = html.Ul(rows, className="mb-0")
            project_options = [{"label": r.name, "value": r.id} for r in df.itertuples()]
            debug_txt = ' | '.join(f"{r.id}:{r.name}" for r in df.itertuples())
        tree_fig = safe_tree_figure()
        return list_children, feedback, tree_fig, project_options, project_options, debug_txt
    except Exception as e:
        return dash.no_update, dbc.Alert(f"Error: {e}", color='danger'), dash.no_update, dash.no_update, dash.no_update, dash.no_update

@app.callback(
    Output('project-name-input','value'),
    Output('project-desc-input','value'),
    Output('sales-tax-input','value'),
    Output('installation-rate-input','value'),
    Output('include-installation-input','value'),
    Output('include-sales-tax-input','value'),
    Input('project-edit-dropdown','value'),
    prevent_initial_call=True
)
def load_project_for_edit(project_id):
    if not project_id:
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('SELECT * FROM projects WHERE id=?', conn, params=(project_id,))
    conn.close()
    if df.empty:
        raise PreventUpdate
    r = df.iloc[0]
    return r['name'], r.get('description',''), round((r.get('sales_tax_rate') or 0)*100,4), round((r.get('installation_rate') or 0)*100,4), ([1] if r.get('include_installation') else []), ([1] if r.get('include_sales_tax') else [])

@app.callback(
    Output('project-create-feedback','children', allow_duplicate=True),
    Output('project-tree','figure', allow_duplicate=True),
    Input('update-project-btn','n_clicks'),
    State('project-edit-dropdown','value'),
    State('project-name-input','value'),
    State('project-desc-input','value'),
    State('sales-tax-input','value'),
    State('installation-rate-input','value'),
    State('include-installation-input','value'),
    State('include-sales-tax-input','value'),
    prevent_initial_call=True
)
def update_project(n_clicks, project_id, name, desc, sales_tax, install_rate, include_install_values, include_tax_values):
    if not n_clicks:
        raise PreventUpdate
    if not project_id or not name:
        return dbc.Alert('Select project and ensure name present', color='danger'), dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    cur.execute('''UPDATE projects SET name=?, description=?, sales_tax_rate=?, installation_rate=?, include_installation=?, include_sales_tax=?, last_modified=CURRENT_TIMESTAMP WHERE id=?''', (
        name.strip(), desc or '', float(sales_tax or 0)/100.0, float(install_rate or 0)/100.0,
        1 if (include_install_values and 1 in include_install_values) else 0,
        1 if (include_tax_values and 1 in include_tax_values) else 0,
        project_id
    ))
    conn.commit(); conn.close()
    return dbc.Alert('Project updated', color='success'), safe_tree_figure()

# Unified refresh for project-edit-dropdown and debug list
## removed refresh_project_dropdown to simplify; debug string handled in create/delete callbacks

# ------------------ Building Management & Sign Assignment ------------------ #
@app.callback(
    Output('building-dropdown', 'options'),
    Output('building-dropdown', 'value'),
    Output('sign-type-dropdown', 'options'),
    Input('assign-project-dropdown', 'value')
)
def load_buildings_for_project(project_id):
    if not project_id:
        return [], None, []
    conn = sqlite3.connect(DATABASE_PATH)
    buildings = pd.read_sql_query("SELECT id, name FROM buildings WHERE project_id = ? ORDER BY id", conn, params=(project_id,))
    sign_types = pd.read_sql_query("SELECT id, name, unit_price FROM sign_types ORDER BY name", conn)
    conn.close()
    building_options = [{"label": r.name, "value": r.id} for r in buildings.itertuples()]
    sign_type_options = [{"label": f"{r.name} (${r.unit_price})", "value": r.id} for r in sign_types.itertuples()]
    return building_options, (building_options[0]['value'] if building_options else None), sign_type_options

@app.callback(
    Output('building-dropdown', 'options', allow_duplicate=True),
    Output('building-action-feedback', 'children', allow_duplicate=True),
    Output('project-tree', 'figure', allow_duplicate=True),
    Input('add-building-btn', 'n_clicks'),
    State('assign-project-dropdown', 'value'),
    State('new-building-name', 'value'),
    State('new-building-desc', 'value'),
    prevent_initial_call=True
)
def add_building(n_clicks, project_id, name, desc):
    if not n_clicks:
        raise PreventUpdate
    if not project_id or not name:
        return dash.no_update, "Select project and enter name", dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    # Duplicate name check (case-insensitive) within project
    cur.execute("SELECT 1 FROM buildings WHERE project_id=? AND LOWER(name)=LOWER(?)", (project_id, name.strip()))
    if cur.fetchone():
        conn.close()
        return dash.no_update, f"Building name '{name}' already exists", dash.no_update
    cur.execute("INSERT INTO buildings (project_id, name, description) VALUES (?,?,?)", (project_id, name.strip(), desc or ''))
    conn.commit()
    buildings = pd.read_sql_query("SELECT id, name FROM buildings WHERE project_id = ? ORDER BY id", conn, params=(project_id,))
    conn.close()
    options = [{"label": r.name, "value": r.id} for r in buildings.itertuples()]
    tree_fig = safe_tree_figure()
    return options, f"Building '{name}' added", tree_fig

@app.callback(
    Output('building-dropdown','options', allow_duplicate=True),
    Output('building-action-feedback','children', allow_duplicate=True),
    Output('project-tree','figure', allow_duplicate=True),
    Input('rename-building-btn','n_clicks'),
    State('assign-project-dropdown','value'),
    State('building-dropdown','value'),
    State('rename-building-input','value'),
    prevent_initial_call=True
)
def rename_building(n_clicks, project_id, building_id, new_name):
    if not n_clicks:
        raise PreventUpdate
    if not (project_id and building_id and new_name and new_name.strip()):
        return dash.no_update, 'Provide building and new name', dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    # uniqueness within project
    cur.execute('SELECT 1 FROM buildings WHERE project_id=? AND LOWER(name)=LOWER(?) AND id<>?', (project_id, new_name.strip(), building_id))
    if cur.fetchone():
        conn.close()
        return dash.no_update, f"Name '{new_name}' already exists", dash.no_update
    cur.execute('UPDATE buildings SET name=?, last_modified=CURRENT_TIMESTAMP WHERE id=?', (new_name.strip(), building_id))
    conn.commit()
    bdf = pd.read_sql_query('SELECT id, name FROM buildings WHERE project_id=? ORDER BY id', conn, params=(project_id,))
    conn.close()
    options = [{'label': r.name, 'value': r.id} for r in bdf.itertuples()]
    return options, 'Building renamed', safe_tree_figure()

def _fetch_building_signs(building_id):
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('''
        SELECT st.name as sign_name, bs.quantity, st.unit_price, (bs.quantity * st.unit_price) as total
        FROM building_signs bs
        JOIN sign_types st ON bs.sign_type_id = st.id
        WHERE bs.building_id = ?
        ORDER BY st.name
    ''', conn, params=(building_id,))
    conn.close()
    return df.to_dict('records')

def _fetch_building_name(building_id):
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        df = pd.read_sql_query('SELECT name FROM buildings WHERE id=?', conn, params=(building_id,))
        conn.close()
        if df.empty:
            return ''
        return df.iloc[0]['name']
    except Exception:
        return ''

@app.callback(
    Output('building-signs-table', 'data'),
    Output('building-action-feedback', 'children', allow_duplicate=True),
    Output('project-tree', 'figure', allow_duplicate=True),
    Input('building-dropdown', 'value'),
    Input('add-sign-to-building-btn', 'n_clicks'),
    Input('save-building-signs-btn', 'n_clicks'),
    State('sign-type-dropdown', 'value'),
    State('sign-qty-input', 'value'),
    State('building-signs-table', 'data'),
    prevent_initial_call=True
)
def manage_building_signs(building_id, add_clicks, save_clicks, sign_type_id, qty, current_rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if not building_id:
        return [], dash.no_update, dash.no_update
    action_msg = dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    if 'add-sign-to-building-btn' in triggered and sign_type_id:
        qty = max(1, int(qty or 1))
        cur.execute("SELECT id, quantity FROM building_signs WHERE building_id=? AND sign_type_id=?", (building_id, sign_type_id))
        existing = cur.fetchone()
        if existing:
            cur.execute("UPDATE building_signs SET quantity=? WHERE id=?", (qty, existing[0]))
        else:
            cur.execute("INSERT INTO building_signs (building_id, sign_type_id, quantity) VALUES (?,?,?)", (building_id, sign_type_id, qty))
        action_msg = "Sign added/updated"
    elif 'save-building-signs-btn' in triggered and current_rows:
        for row in current_rows:
            name = row.get('sign_name')
            q = max(0, int(row.get('quantity') or 0))
            cur.execute("SELECT id FROM sign_types WHERE name = ?", (name,))
            st_row = cur.fetchone()
            if not st_row:
                continue
            st_id = st_row[0]
            cur.execute("SELECT id FROM building_signs WHERE building_id=? AND sign_type_id=?", (building_id, st_id))
            ex = cur.fetchone()
            if ex:
                cur.execute("UPDATE building_signs SET quantity=? WHERE id=?", (q, ex[0]))
            else:
                cur.execute("INSERT INTO building_signs (building_id, sign_type_id, quantity) VALUES (?,?,?)", (building_id, st_id, q))
        action_msg = "Quantities saved"
    if action_msg is not dash.no_update:
        conn.commit()
    conn.close()
    data = _fetch_building_signs(building_id)
    tree_fig = safe_tree_figure()
    return data, action_msg, tree_fig

# -------- Sign Groups within Projects Tab -------- #
@app.callback(
    Output('project-group-dropdown','options'),
    Input('assign-project-dropdown','value')
)
def load_group_options_for_project(_project_id):
    conn = sqlite3.connect(DATABASE_PATH)
    gdf = pd.read_sql_query('SELECT id, name FROM sign_groups ORDER BY name', conn)
    conn.close()
    return ([{'label': r.name, 'value': r.id} for r in gdf.itertuples()]) if not gdf.empty else []

def _fetch_building_groups(building_id):
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('''SELECT sg.name as group_name, bsg.quantity, sg.id as group_id
                               FROM building_sign_groups bsg
                               JOIN sign_groups sg ON bsg.group_id = sg.id
                               WHERE bsg.building_id=? ORDER BY sg.name''', conn, params=(building_id,))
    conn.close()
    return df.to_dict('records')

def _fetch_assigned_group_options(building_id):
    """Return dropdown options for groups already assigned to a building."""
    if not building_id:
        return []
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('''SELECT sg.id, sg.name FROM sign_groups sg
                               JOIN building_sign_groups bsg ON bsg.group_id=sg.id
                               WHERE bsg.building_id=? ORDER BY sg.name''', conn, params=(building_id,))
    conn.close()
    return [{'label': r['name'], 'value': r['id']} for _, r in df.iterrows()]

@app.callback(
    Output('project-building-groups-table','data'),
    Output('building-groups-feedback','children', allow_duplicate=True),
    Output('project-tree','figure', allow_duplicate=True),
    Input('building-dropdown','value'),
    Input('add-group-to-building-btn','n_clicks'),
    Input('save-building-groups-btn','n_clicks'),
    State('project-group-dropdown','value'),
    State('group-qty-input','value'),
    State('project-building-groups-table','data'),
    prevent_initial_call=True
)
def manage_building_groups_inline(building_id, add_clicks, save_clicks, group_id, qty, rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if not building_id:
        return [], dash.no_update, dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    msg = dash.no_update
    if 'add-group-to-building-btn' in triggered and group_id:
        q = max(1, int(qty or 1))
        cur.execute('SELECT id, quantity FROM building_sign_groups WHERE building_id=? AND group_id=?', (building_id, group_id))
        ex = cur.fetchone()
        if ex:
            cur.execute('UPDATE building_sign_groups SET quantity=? WHERE id=?', (q, ex[0]))
        else:
            cur.execute('INSERT INTO building_sign_groups (building_id, group_id, quantity) VALUES (?,?,?)', (building_id, group_id, q))
        msg = 'Group added/updated'
    elif 'save-building-groups-btn' in triggered and rows:
        for r in rows:
            gname = r.get('group_name')
            q = max(0, int(r.get('quantity') or 0))
            cur.execute('SELECT id FROM sign_groups WHERE name=?', (gname,))
            gr = cur.fetchone()
            if not gr:
                continue
            cur.execute('SELECT id FROM building_sign_groups WHERE building_id=? AND group_id=?', (building_id, gr[0]))
            ex = cur.fetchone()
            if ex:
                cur.execute('UPDATE building_sign_groups SET quantity=? WHERE id=?', (q, ex[0]))
        msg = 'Group quantities saved'
    conn.commit()
    group_rows = _fetch_building_groups(building_id)
    conn.close()
    tree_fig = safe_tree_figure()
    return group_rows, msg, tree_fig

# --- Filter group options (Show All vs project-assigned) & populate deletion dropdown --- #
@app.callback(
    Output('project-group-dropdown','options', allow_duplicate=True),
    Output('assigned-group-delete-dropdown','options'),
    Input('assign-project-dropdown','value'),
    Input('building-dropdown','value'),
    Input('project-groups-show-all','value'),
    Input('project-building-groups-table','data'),  # refresh after add/save quantity
    prevent_initial_call=True
)
def refresh_project_group_dropdowns(project_id, building_id, show_all_values, _group_table_rows):
    try:
        # Deletion dropdown
        assigned_opts = _fetch_assigned_group_options(building_id)
        # Group options
        show_all = bool(show_all_values and 'all' in show_all_values)
        if show_all:
            conn = sqlite3.connect(DATABASE_PATH)
            gdf = pd.read_sql_query('SELECT id, name FROM sign_groups ORDER BY name', conn)
            conn.close()
            group_opts = [{'label': r['name'], 'value': r['id']} for _, r in gdf.iterrows()]
        else:
            if project_id:
                conn = sqlite3.connect(DATABASE_PATH)
                gdf = pd.read_sql_query('''SELECT DISTINCT sg.id, sg.name FROM sign_groups sg
                                            JOIN building_sign_groups bsg ON bsg.group_id=sg.id
                                            JOIN buildings b ON bsg.building_id=b.id
                                            WHERE b.project_id=? ORDER BY sg.name''', conn, params=(project_id,))
                conn.close()
                group_opts = [{'label': r['name'], 'value': r['id']} for _, r in gdf.iterrows()]
            else:
                group_opts = []
        return group_opts, assigned_opts
    except Exception as e:
        print(f"[groups][filter][error] {e}")
        return dash.no_update, dash.no_update

# --- Remove group from building --- #
@app.callback(
    Output('project-building-groups-table','data', allow_duplicate=True),
    Output('building-groups-feedback','children', allow_duplicate=True),
    Output('project-tree','figure', allow_duplicate=True),
    Output('assigned-group-delete-dropdown','options', allow_duplicate=True),
    Input('remove-group-from-building-btn','n_clicks'),
    State('building-dropdown','value'),
    State('assigned-group-delete-dropdown','value'),
    prevent_initial_call=True
)
def remove_group_from_building(n_clicks, building_id, group_id):
    if not n_clicks:
        raise PreventUpdate
    if not building_id or not group_id:
        return dash.no_update, 'Select building and group to remove', dash.no_update, dash.no_update
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        cur.execute('DELETE FROM building_sign_groups WHERE building_id=? AND group_id=?', (building_id, group_id))
        conn.commit()
        df = pd.read_sql_query('''SELECT sg.name as group_name, bsg.quantity, sg.id as group_id
                                   FROM building_sign_groups bsg
                                   JOIN sign_groups sg ON bsg.group_id=sg.id
                                   WHERE bsg.building_id=? ORDER BY sg.name''', conn, params=(building_id,))
        conn.close()
        table_rows = df.to_dict('records')
        assigned_opts = _fetch_assigned_group_options(building_id)
        return table_rows, 'Group removed', safe_tree_figure(), assigned_opts
    except Exception as e:
        print(f"[groups][remove][error] {e}")
        return dash.no_update, f'Error removing: {e}', dash.no_update, dash.no_update

# --- Global error toast trigger (simple heuristic scanning for keyword) --- #
@app.callback(
    Output('app-error-toast','is_open'),
    Output('app-error-toast','children'),
    Input('building-groups-feedback','children'),
    Input('project-create-feedback','children'),
    prevent_initial_call=True
)
def surface_errors(*feedback_children):
    for child in feedback_children:
        if not child:
            continue
        # Convert component to string for heuristic search
        text = ''
        try:
            if isinstance(child, (str, int, float)):
                text = str(child)
            elif hasattr(child, 'props') and 'children' in child.props:
                # For dbc.Alert etc.
                inner = child.props.get('children')
                text = inner if isinstance(inner, str) else str(inner)
            else:
                text = str(child)
        except Exception:
            text = str(child)
        lowered = text.lower()
        if 'error' in lowered or 'warning' in lowered or 'already exists' in lowered:
            return True, text
    raise PreventUpdate

# ------------------ Runtime status (DB & Code refresh timestamps) ------------------ #
def _humanize_delta(delta_seconds: float) -> str:
    try:
        if delta_seconds < 0:
            delta_seconds = 0
        mins = delta_seconds / 60.0
        if mins < 1:
            return f"{int(delta_seconds)}s ago"
        if mins < 60:
            return f"{int(mins)}m ago"
        hours = mins / 60.0
        if hours < 24:
            return f"{int(hours)}h ago"
        days = hours / 24.0
        if days < 7:
            return f"{int(days)}d ago"
        weeks = days / 7.0
        if weeks < 4:
            return f"{int(weeks)}w ago"
        months = days / 30.0
        if months < 12:
            return f"{int(months)}mo ago"
        years = days / 365.0
        return f"{years:.1f}y ago"
    except Exception:
        return "n/a"

def _get_db_timestamp() -> tuple[str, str]:
    try:
        if DATABASE_PATH and os.path.exists(DATABASE_PATH):
            ts = os.path.getmtime(DATABASE_PATH)
            dt = datetime.fromtimestamp(ts, timezone.utc)
            rel = _humanize_delta((datetime.now(timezone.utc) - dt).total_seconds())
            return rel, dt.isoformat() + 'Z'
    except Exception:
        pass
    return 'unknown', ''

def _get_code_timestamp() -> tuple[str, str]:
    # Preferred: deployment_info.json (deployed_at)
    try:
        deployment_file = Path('deployment_info.json')
        if deployment_file.exists():
            import json as _json
            try:
                info = _json.loads(deployment_file.read_text())
                deployed_at = info.get('deployed_at') or info.get('timestamp')
                if deployed_at:
                    # Normalize Z
                    ts_txt = deployed_at.replace('Z','')
                    try:
                        dt = datetime.fromisoformat(ts_txt)
                    except Exception:
                        dt = datetime.now(timezone.utc)
                    # Assume provided timestamp is UTC
                    rel = _humanize_delta((datetime.now(timezone.utc) - dt).total_seconds())
                    return rel, (dt.isoformat() + 'Z')
            except Exception:
                pass
        # Fallback: max mtime of a small set of source files/dirs
        candidates = []
        for relp in ['app.py', 'utils', 'scripts']:
            p = Path(relp)
            if p.is_file():
                candidates.append(p.stat().st_mtime)
            elif p.is_dir():
                # Shallow scan only for performance
                for child in p.iterdir():
                    if child.suffix in ('.py', '.json') and child.is_file():
                        candidates.append(child.stat().st_mtime)
        if candidates:
            ts = max(candidates)
            dt = datetime.fromtimestamp(ts, timezone.utc)
            rel = _humanize_delta((datetime.now(timezone.utc) - dt).total_seconds())
            return rel, dt.isoformat() + 'Z'
    except Exception:
        pass
    return 'unknown', ''

@app.callback(
    Output('runtime-status','children'),
    Input('status-refresh-interval','n_intervals'),
    prevent_initial_call=False
)
def update_runtime_status(_n):
    """Unified footer status (Python version, DB size & age, code age, optional SVG short status).

    Hides entirely if SIGN_APP_HIDE_ENV_NOTICE=1.
    """
    if os.getenv('SIGN_APP_HIDE_ENV_NOTICE','').lower() in ('1','true','yes'):
        return ''
    try:
        db_rel, db_iso = _get_db_timestamp()
        code_rel, code_iso = _get_code_timestamp()
        py_ver = f"Py {sys.version_info.major}.{sys.version_info.minor}"
        # DB size
        size_txt = ''
        try:
            if DATABASE_PATH and os.path.exists(DATABASE_PATH):
                sz = os.path.getsize(DATABASE_PATH)
                if sz >= 1_000_000:
                    size_txt = f"{(sz/1_000_000):.1f}MB"
                elif sz >= 10_000:
                    size_txt = f"{(sz/1000):.0f}KB"
                else:
                    size_txt = f"{sz}B"
        except Exception:
            pass
        svg_status = os.environ.get('SIGN_APP_SVG_STATUS')
        svg_short = ''
        if svg_status and svg_status not in {'ok'}:
            svg_short = 'SVG ' + svg_status.split(':',1)[0]
        parts = [py_ver]
        if size_txt:
            parts.append(f"DB {size_txt} ({db_rel})")
        else:
            parts.append(f"DB {db_rel}")
        parts.append(f"Code {code_rel}")
        if svg_short:
            parts.append(svg_short)
        # Keep ISO timestamps as title tooltip
        title_attr = None
        if db_iso or code_iso:
            title_attr = f"DB: {db_iso or 'n/a'} | Code: {code_iso or 'n/a'}"
        return html.Small(' • '.join(parts), className='text-muted', title=title_attr)
    except Exception as e:
        return html.Small(f"Status unavailable: {e}", className='text-muted')

# ------------------ Estimate Generation & Export ------------------ #
@app.callback(
    Output('estimate-table', 'data'),
    Output('estimate-summary', 'children'),
    Output('export-estimate-btn', 'disabled'),
    Output('export-estimate-pdf-btn', 'disabled'),
    Input('generate-estimate-btn', 'n_clicks'),
    State('estimate-project-dropdown', 'value'),
    State('estimate-building-dropdown', 'value'),
    State('price-calc-mode','value'),
    State('install-mode','value'),
    State('install-percent-input','value'),
    State('install-per-sign-rate','value'),
    State('install-per-area-rate','value'),
    State('install-hours-input','value'),
    State('install-hourly-rate-input','value'),
    State('auto-install-use','value'),
    State('embed-images-store','data'),
    State('exterior-only-toggle','value'),
    State('non-exterior-only-toggle','value'),
    prevent_initial_call=True
)
def generate_estimate(n_clicks, project_id, building_id, price_mode, install_mode, inst_percent, inst_per_sign, inst_per_area, inst_hours, inst_hourly, auto_install_toggle, embed_store, exterior_toggle, non_exterior_toggle):
    if not n_clicks:
        raise PreventUpdate
    if db_manager is None:
        return [], dbc.Alert("Database not initialized", color='danger'), True
    # building_id may now be list (multi select). Normalize.
    building_ids = []
    if isinstance(building_id, list):
        building_ids = [b for b in building_id if b is not None]
    elif building_id:
        building_ids = [building_id]
    if not project_id and not building_ids:
        return [], dbc.Alert("Select a project or building(s)", color='warning'), True
    # If only buildings chosen, derive project (assume same project; take first)
    if building_ids and not project_id:
        conn = sqlite3.connect(DATABASE_PATH)
        placeholders = ','.join(['?']*len(building_ids))
        pdf = pd.read_sql_query(f'SELECT DISTINCT project_id FROM buildings WHERE id IN ({placeholders})', conn, params=tuple(building_ids))
        conn.close()
        if not pdf.empty:
            project_id = pdf.iloc[0]['project_id']
    def _coerce(v):
        try: return float(v or 0)
        except: return 0.0
    inst_percent = _coerce(inst_percent)
    inst_per_sign = _coerce(inst_per_sign)
    inst_per_area = _coerce(inst_per_area)
    inst_hours = _coerce(inst_hours)
    inst_hourly = _coerce(inst_hourly)

    use_default = (price_mode == 'per_sign' and install_mode == 'percent')
    auto_enabled = bool(auto_install_toggle and 1 in auto_install_toggle)
    meta = {}
    if use_default:
        estimate_data = db_manager.get_project_estimate(project_id) or []
    else:
        from utils.estimate_core import compute_custom_estimate
        # Provide building filter directly if user selected building_ids; else None for all
        estimate_result = compute_custom_estimate(
            DATABASE_PATH, project_id, building_ids if building_ids else None,
            price_mode, install_mode,
            inst_percent, inst_per_sign, inst_per_area, inst_hours, inst_hourly,
            auto_enabled, return_meta=True
        )
        if isinstance(estimate_result, tuple):
            estimate_data, meta = estimate_result
        else:
            estimate_data = estimate_result
    if not estimate_data:
        return [], dbc.Alert("No data", color='warning'), True
    # If we used default path and building_ids provided, filter manually
    if use_default and building_ids:
        conn = sqlite3.connect(DATABASE_PATH)
        placeholders = ','.join(['?']*len(building_ids))
        ndf = pd.read_sql_query(f'SELECT id, name FROM buildings WHERE id IN ({placeholders})', conn, params=tuple(building_ids))
        conn.close()
        selected_names = set(ndf['name'].tolist())
        estimate_data = [r for r in estimate_data if r['Building'] in selected_names or r['Building']=='ALL']
    if not estimate_data:
        return [], dbc.Alert("No data for selection", color='warning'), True
    df = pd.DataFrame(estimate_data)
    # Exterior-only filter: install_type source needed. Join sign_types to determine classification.
    # Harmonize toggles (Dash passes them as lists)
    ext_only = exterior_toggle and 'ext_only' in exterior_toggle
    non_ext_only = non_exterior_toggle and 'non_ext_only' in non_exterior_toggle
    # If both toggles somehow enabled, prioritize no filter (user conflict); could also choose to raise-warning
    if ext_only and non_ext_only:
        ext_only = non_ext_only = False
    exterior_filtered = False
    non_exterior_filtered = False
    if ext_only or non_ext_only:
        try:
            conn = sqlite3.connect(DATABASE_PATH)
            it_map_df = pd.read_sql_query('SELECT name, install_type FROM sign_types', conn)
            conn.close()
            it_map = {r['name'].lower(): (r['install_type'] or '') for _, r in it_map_df.iterrows()}
            def _is_ext(item):
                base = (str(item).split('Group:')[-1].strip()).lower()
                return 'ext' in (it_map.get(base, '') or '').lower()
            if ext_only:
                df = df[[ _is_ext(it) for it in df['Item'] ]]
                exterior_filtered = True
            elif non_ext_only:
                df = df[[ (not _is_ext(it)) for it in df['Item'] ]]
                non_exterior_filtered = True
        except Exception as _fe:
            print(f"[estimate][ext-only][warn] {_fe}")
    total = df['Total'].sum() if 'Total' in df else 0
    # Build chips & meta display
    chips = []
    chips.append(dbc.Badge(f"Total: ${total:,.2f}", color='primary', className='me-1'))
    if building_ids:
        chips.append(dbc.Badge(f"Buildings: {len(building_ids)}", color='secondary', className='me-1'))
    if not use_default and meta:
        chips.append(dbc.Badge(f"Signs: {int(meta.get('total_sign_count',0))}", color='info', className='me-1'))
        if meta.get('total_area'):
            chips.append(dbc.Badge(f"Area: {meta['total_area']:.1f} sq ft", color='light', text_color='dark', className='me-1'))
        if meta.get('auto_install_amount_per_sign'):
            chips.append(dbc.Badge(f"Auto Inst $/sign sum: ${meta['auto_install_amount_per_sign']:.2f}", color='warning', className='me-1'))
        if meta.get('auto_install_hours'):
            chips.append(dbc.Badge(f"Auto Inst Hours: {meta['auto_install_hours']:.1f}", color='warning', className='me-1'))
        if meta.get('install_cost'):
            chips.append(dbc.Badge(f"Install: ${meta['install_cost']:.2f}", color='danger', className='me-1'))
    # Pricing mode note
    note_lines = []
    if price_mode == 'per_area':
        note_lines.append("Pricing basis: Area * Material Rate")
    if not use_default:
        note_lines.append(f"Install mode: {install_mode}")
    if exterior_filtered:
        chips.append(dbc.Badge("Filtered: Exterior Only", color='dark', className='me-1'))
    if non_exterior_filtered:
        chips.append(dbc.Badge("Filtered: Non-Exterior Only", color='dark', className='me-1'))
    summary = html.Div([
        html.Div(chips, className='mb-1'),
        html.Small(" | ".join(note_lines)) if note_lines else None
    ])
    return df.to_dict('records'), summary, False, False

@app.callback(
    Output('estimate-download', 'data'),
    Input('export-estimate-btn', 'n_clicks'),
    State('estimate-project-dropdown', 'value'),
    State('estimate-building-dropdown', 'value'),
    State('price-calc-mode','value'),
    State('install-mode','value'),
    State('install-percent-input','value'),
    State('install-per-sign-rate','value'),
    State('install-per-area-rate','value'),
    State('install-hours-input','value'),
    State('install-hourly-rate-input','value'),
    State('auto-install-use','value'),
    prevent_initial_call=True
)
def export_estimate(n_clicks, project_id, building_id, price_mode, install_mode, inst_percent, inst_per_sign, inst_per_area, inst_hours, inst_hourly, auto_install_toggle, embed_store):
    if not n_clicks:
        raise PreventUpdate
    # Normalize multi select
    building_ids = []
    if isinstance(building_id, list):
        building_ids = [b for b in building_id if b is not None]
    elif building_id:
        building_ids = [building_id]
    if not project_id and not building_ids:
        return dash.no_update
    if db_manager is None:
        return dash.no_update
    # Derive project id from building if needed
    if building_ids and not project_id:
        conn = sqlite3.connect(DATABASE_PATH)
        placeholders = ','.join(['?']*len(building_ids))
        pdf = pd.read_sql_query(f'SELECT DISTINCT project_id FROM buildings WHERE id IN ({placeholders})', conn, params=tuple(building_ids))
        conn.close()
        if not pdf.empty:
            project_id = pdf.iloc[0]['project_id']
    try:
        # Reuse generate_estimate core logic by lightweight inline recompute (duplicated minimal branch for export)
        def _coerce(v):
            try: return float(v or 0)
            except: return 0.0
        inst_percent=_coerce(inst_percent); inst_per_sign=_coerce(inst_per_sign); inst_per_area=_coerce(inst_per_area); inst_hours=_coerce(inst_hours); inst_hourly=_coerce(inst_hourly)
        use_default = (price_mode=='per_sign' and install_mode=='percent')
        if use_default:
            estimate_data = db_manager.get_project_estimate(project_id) or []
        else:
            estimate_data = []
            conn = sqlite3.connect(DATABASE_PATH)
            proj_df = pd.read_sql_query('SELECT * FROM projects WHERE id=?', conn, params=(project_id,))
            if proj_df.empty:
                conn.close(); return dash.no_update
            project = proj_df.iloc[0]
            buildings = pd.read_sql_query('SELECT * FROM buildings WHERE project_id=?', conn, params=(project_id,))
            grand_subtotal=0.0; total_sign_count=0; total_area=0.0; auto_install_amount_per_sign=0.0; auto_install_hours=0.0
            auto_enabled = bool(auto_install_toggle and 1 in auto_install_toggle)
            for _, b in buildings.iterrows():
                b_sub=0.0
                signs = pd.read_sql_query('''SELECT st.name, st.unit_price, st.material, st.width, st.height, st.price_per_sq_ft, st.per_sign_install_rate, st.install_time_hours, bs.quantity
                                              FROM building_signs bs JOIN sign_types st ON bs.sign_type_id=st.id WHERE bs.building_id=?''', conn, params=(b['id'],))
                for _, s in signs.iterrows():
                    qty = s['quantity'] or 0
                    total_sign_count += qty
                    width=_coerce(s['width']); height=_coerce(s['height']); ppsf=_coerce(s['price_per_sq_ft'])
                    area = width*height if width and height else 0
                    total_area += area*qty
                    ps_install=_coerce(s.get('per_sign_install_rate'))
                    if ps_install>0:
                        auto_install_amount_per_sign += ps_install * qty
                    inst_time=_coerce(s.get('install_time_hours'))
                    if inst_time>0:
                        auto_install_hours += inst_time * qty
                    unit_price = compute_unit_price(s.to_dict(), price_mode)
                    line_total = unit_price*qty
                    b_sub += line_total
                    estimate_data.append({'Building':b['name'],'Item':s['name'],'Material':s['material'],'Dimensions':f"{width} x {height}" if width and height else '', 'Quantity':qty,'Unit_Price':unit_price,'Total':line_total})
                groups = pd.read_sql_query('''SELECT sg.id, sg.name, bsg.quantity FROM building_sign_groups bsg JOIN sign_groups sg ON bsg.group_id=sg.id WHERE bsg.building_id=?''', conn, params=(b['id'],))
                for _, g in groups.iterrows():
                    group_members = pd.read_sql_query('''SELECT st.name, st.unit_price, st.width, st.height, st.price_per_sq_ft, st.per_sign_install_rate, st.install_time_hours, st.material_multiplier, sgm.quantity FROM sign_group_members sgm JOIN sign_types st ON sgm.sign_type_id=st.id WHERE sgm.group_id=?''', conn, params=(g['id'],))
                    group_cost_unit=0.0
                    for _, m in group_members.iterrows():
                        m_qty = (m['quantity'] or 0)
                        width=_coerce(m['width']); height=_coerce(m['height']); ppsf=_coerce(m['price_per_sq_ft'])
                        area=width*height if width and height else 0
                        total_sign_count += m_qty * g['quantity']
                        total_area += area * m_qty * g['quantity']
                        ps_install=_coerce(m.get('per_sign_install_rate'))
                        if ps_install>0:
                            auto_install_amount_per_sign += ps_install * m_qty * g['quantity']
                        inst_time=_coerce(m.get('install_time_hours'))
                        if inst_time>0:
                            auto_install_hours += inst_time * m_qty * g['quantity']
                        m_unit = compute_unit_price(m.to_dict(), price_mode)
                        group_cost_unit += m_unit * m_qty
                    unit_price=group_cost_unit; line_total=unit_price * g['quantity']; b_sub+=line_total
                    estimate_data.append({'Building':b['name'],'Item':f"Group: {g['name']}",'Material':'Various','Dimensions':'','Quantity':g['quantity'],'Unit_Price':unit_price,'Total':line_total})
                grand_subtotal+=b_sub
            install_cost = compute_install_cost(
                install_mode, grand_subtotal, total_sign_count, total_area,
                inst_percent, inst_per_sign, inst_per_area, inst_hours, inst_hourly,
                auto_enabled, auto_install_amount_per_sign, auto_install_hours
            )
            if install_mode!='none' and install_cost>0:
                estimate_data.append({'Building':'ALL','Item':'Installation','Material':'','Dimensions':'','Quantity':1,'Unit_Price':install_cost,'Total':install_cost})
            if bool(project.get('include_sales_tax')) and project.get('sales_tax_rate'):
                taxable_total = sum(r['Total'] for r in estimate_data if r['Building']!='ALL' or r['Item']=='Installation')
                tax_cost = taxable_total * float(project['sales_tax_rate'])
                estimate_data.append({'Building':'ALL','Item':'Sales Tax','Material':'','Dimensions':'','Quantity':1,'Unit_Price':tax_cost,'Total':tax_cost})
            conn.close()
        if not estimate_data:
            return dash.no_update
        # Filter to building if requested
        if building_ids:
            conn = sqlite3.connect(DATABASE_PATH)
            placeholders = ','.join(['?']*len(building_ids))
            ndf = pd.read_sql_query(f'SELECT name FROM buildings WHERE id IN ({placeholders})', conn, params=tuple(building_ids))
            conn.close()
            selected_names = set(ndf['name'].tolist())
            estimate_data = [r for r in estimate_data if r['Building'] in selected_names or r['Building']=='ALL']
            if not estimate_data:
                return dash.no_update
        df = pd.DataFrame(estimate_data)
        buffer = io.BytesIO()
        from openpyxl.drawing.image import Image as XLImage
        from tempfile import NamedTemporaryFile
        logo_path = Path('assets') / 'LSI_Logo.svg'
        try:
            import cairosvg  # optional dependency
        except Exception:
            cairosvg = None
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Estimate')
            wb = writer.book
            ws = wb['Estimate']
            # Preload image paths map (sign name -> path)
            image_map = {}
            try:
                conn = sqlite3.connect(DATABASE_PATH)
                idf = pd.read_sql_query('SELECT name, image_path FROM sign_types WHERE image_path IS NOT NULL AND image_path<>""', conn)
                conn.close()
                for _, ir in idf.iterrows():
                    ip = ir['image_path']
                    if ip and Path(ip).exists():
                        image_map[ir['name'].lower()] = Path(ip)
            except Exception as e:
                print(f"[excel][thumb-preload][warn] {e}")
            # Insert branding header above table
            ws.insert_rows(1, amount=4)
            ws.merge_cells('A1:D3')
            ws['A1'] = 'Sign Estimation Project Export'
            # Safely copy font/alignment if row 5 exists
            try:
                base_font = ws['A5'].font
                base_align = ws['A5'].alignment
            except Exception:
                base_font = ws['A1'].font
                base_align = ws['A1'].alignment
            ws['A1'].font = base_font.copy(bold=True)
            ws['A1'].alignment = base_align.copy(horizontal='left', vertical='center', wrap_text=True)
            if cairosvg and logo_path.exists():
                try:
                    with NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                        png_temp = tmp.name
                    cairosvg.svg2png(url=str(logo_path), write_to=png_temp, output_width=320)
                    img = XLImage(png_temp)
                    img.anchor = 'F1'
                    ws.add_image(img)
                except Exception:
                    pass
            # Optional thumbnail column (insert after A header if images present)
            embed_images = True
            try:
                embed_images = bool(embed_store and embed_store.get('embed'))
            except Exception:
                embed_images = True
            try:
                if image_map and embed_images:
                    # Add new column for thumbnails at column A shifting existing
                    ws.insert_cols(1, amount=1)
                    ws['A5'] = 'Image'
                    from utils.image_cache import get_or_build_thumbnail
                    max_row = ws.max_row
                    # Data starts at row 5 now (header offset + inserted rows) ; find column indexes for Item and Building to compute matching sign name
                    # Find 'Item' header (shifted one col to right due to insert).
                    header_row = 5
                    item_col = None
                    building_col = None
                    for cell in ws[header_row]:
                        if cell.value == 'Item':
                            item_col = cell.column
                        if cell.value == 'Building':
                            building_col = cell.column
                    if item_col:
                        for r_idx in range(header_row+1, max_row+1):
                            cell_item = ws.cell(row=r_idx, column=item_col).value
                            if not cell_item:
                                continue
                            base_item = str(cell_item)
                            if base_item.startswith('Group:'):
                                base_item = base_item.replace('Group:','').strip()
                            key = base_item.lower()
                            pth = image_map.get(key)
                            if not pth:
                                continue
                            tp = get_or_build_thumbnail(pth, 120, 60)
                            if tp and Path(tp).exists():
                                try:
                                    thumb_img = XLImage(str(tp))
                                    thumb_img.anchor = f"A{r_idx}"
                                    ws.add_image(thumb_img)
                                except Exception:
                                    pass
                    ws.column_dimensions['A'].width = 14
            except Exception as ee:
                print(f"[excel][thumb][warn] {ee}")
            # Footer branding line after table
            footer_row = ws.max_row + 2
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=6)
            ws.cell(row=footer_row, column=1, value='© 2025 LSI Graphics, LLC — Generated by Sign Package Estimator')
            try:
                ws.cell(row=footer_row, column=1).font = base_font.copy(italic=True, size=11)
            except Exception:
                pass
            ws.cell(row=footer_row, column=1).alignment = base_align.copy(horizontal='center')
        buffer.seek(0)
        suffix = ''
        if building_ids:
            suffix = f"_buildings_{'-'.join(map(str, building_ids))}"
        return dict(content=base64.b64encode(buffer.read()).decode(), filename=f'project_{project_id}{suffix}_estimate.xlsx', type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"[export][error] {e}")
        err_buf = io.BytesIO()
        with pd.ExcelWriter(err_buf, engine='openpyxl') as writer:
            pd.DataFrame([{"Error": str(e)}]).to_excel(writer, index=False, sheet_name='Error')
        err_buf.seek(0)
        return dict(content=base64.b64encode(err_buf.read()).decode(), filename=f'project_{project_id}_export_error.xlsx', type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ------------------ PDF Export ------------------ #
@app.callback(
    Output('estimate-pdf-download','data'),
    Input('export-estimate-pdf-btn','n_clicks'),
    State('estimate-table','data'),
    State('estimate-summary','children'),
    State('estimate-pdf-title','value'),
    State('disable-logo-toggle','value'),
    State('disable-logo-store','data'),
    prevent_initial_call=True
)
def export_estimate_pdf(n_clicks, table_data, summary_children, pdf_title, disable_logo_toggle, disable_logo_store):
    if not n_clicks:
        raise PreventUpdate
    if not table_data:
        raise PreventUpdate
    try:
        def _extract_text(node):
            # Handle Dash components (dict-like) extracting 'children'
            if node is None:
                return ''
            # Dash components may appear as dict with 'props'
            if isinstance(node, dict):
                props = node.get('props') or {}
                ch = props.get('children')
                if isinstance(ch, (list, tuple)):
                    return ' '.join(filter(None, (_extract_text(c) for c in ch)))
                return _extract_text(ch)
            # List/tuple of nodes
            if isinstance(node, (list, tuple)):
                return ' '.join(filter(None, (_extract_text(c) for c in node)))
            # Primitive
            return str(node)
        summary_text = _extract_text(summary_children).strip()
        # Collapse excessive whitespace
        import re as _re
        summary_text = _re.sub(r'\s+', ' ', summary_text)
        from utils.pdf_export import generate_estimate_pdf
        from dash import dcc
        import os as _os
        # precedence: explicit toggle > store > env var
        if disable_logo_toggle:
            disable_logo = 'no_logo' in disable_logo_toggle
        elif disable_logo_store:
            try:
                disable_logo = bool(disable_logo_store.get('disable'))
            except Exception:
                disable_logo = False
        else:
            disable_logo = bool(_os.environ.get('ESTIMATE_PDF_DISABLE_LOGO'))
        # Build multi-image lookup: sign_name -> ordered list of image paths (cover first)
        multi_lookup = {}
        try:
            import sqlite3 as _sqlite
            conn = _sqlite.connect(DATABASE_PATH)
            cur = conn.cursor()
            # For all distinct sign Items in the table (strip any 'Group:' prefix and quantity suffix)
            item_names = set()
            for r in table_data:
                raw_item = str(r.get('Item') or '').strip()
                if not raw_item:
                    continue
                base_item = raw_item.split('Group:')[-1].strip() if raw_item.startswith('Group:') else raw_item
                # Remove trailing quantity pattern e.g., "(3)" only if at end
                import re as _re2
                cleaned = _re2.sub(r'\(\d+\)$','', base_item).strip()
                if cleaned:
                    item_names.add(cleaned)
            # Query once per name (could optimize with join) preserving display_order
            for name in item_names:
                try:
                    cur.execute('''SELECT sti.image_path FROM sign_type_images sti 
                                   JOIN sign_types st ON st.id=sti.sign_type_id 
                                   WHERE st.name=? ORDER BY sti.display_order ASC, sti.id ASC''', (name,))
                    rows = [r[0] for r in cur.fetchall() if r and r[0]]
                    existing = [p for p in rows if os.path.exists(p)]
                    if existing:
                        multi_lookup[name] = existing
                except Exception:
                    continue
            conn.close()
        except Exception as _mle:
            print(f"[pdf][multi-image][warn] {_mle}")
        pdf_bytes, diag = generate_estimate_pdf(table_data, summary_text, pdf_title, str(DATABASE_PATH), disable_logo=disable_logo, multi_image_lookup=multi_lookup)
        sig = diag.get('head_signature')
        eof_present = diag.get('eof_present')
        logo_diag = diag.get('logo', {})
        print(f"[pdf][diag] size={diag.get('size')} sig={sig!r} eof={eof_present} sha1={diag.get('sha1')[:10]} rows={diag.get('row_count')} ext={diag.get('exterior_count')} int={diag.get('interior_count')} logo={logo_diag} embed={diag.get('embed_images')} appendix={diag.get('appendix_count')}")
        if not (isinstance(sig, (bytes, bytearray)) and sig.startswith(b'%PDF') and eof_present):
            print('[pdf][diag][warn] signature or EOF missing, raising for fallback')
            raise ValueError('Generated PDF failed validation')
        filename = (pdf_title.strip().replace(' ','_') if pdf_title else 'estimate') + '.pdf'
        # Use binary-safe send_bytes to avoid any potential corruption on large files or special characters
        return dcc.send_bytes(lambda buff: buff.write(pdf_bytes), filename)
    except Exception as e:
        print(f'[export-pdf][error] {e}')
        # Build a minimal fallback PDF with just error notice
        try:
            from reportlab.lib.pagesizes import LETTER
            from reportlab.pdfgen import canvas
            err_buf=io.BytesIO(); c=canvas.Canvas(err_buf, pagesize=LETTER)
            c.setFont('Helvetica',12); c.drawString(72, LETTER[1]-100, 'Estimate PDF Generation Failed.')
            c.setFont('Helvetica',10); c.drawString(72, LETTER[1]-120, f'Error: {str(e)[:140]}')
            c.drawString(72, LETTER[1]-140, 'Please contact support or retry export.')
            c.showPage(); c.save(); err_buf.seek(0)
            from dash import dcc
            fb_bytes = err_buf.read()
            return dcc.send_bytes(lambda buff: buff.write(fb_bytes), 'estimate_error.pdf')
        except Exception as ee:
            print(f'[export-pdf][fallback-error] {ee}')
            raise PreventUpdate

# ------------------ Tree PNG Export ------------------ #
@app.callback(
    Output('tree-png-download','data'),
    Input('export-tree-png-btn','n_clicks'),
    State('project-tree','figure'),
    prevent_initial_call=True
)
def export_tree_png(n_clicks, fig_dict):
    if not n_clicks:
        raise PreventUpdate
    try:
        import plotly.graph_objects as go
        from PIL import Image as PILImage, ImageDraw, ImageFont
        from dash import dcc
        # Validate fig_dict structure
        if not isinstance(fig_dict, dict) or 'data' not in fig_dict:
            # Create error image
            err_img = PILImage.new('RGB', (800, 200), 'white')
            d = ImageDraw.Draw(err_img)
            try:
                f = ImageFont.truetype('Arial.ttf', 16)
            except Exception:
                f = ImageFont.load_default()
            d.text((20, 40), 'Tree figure unavailable (invalid data).', fill='red', font=f)
            print(f"[png][diag] invalid fig_dict keys={list(fig_dict.keys()) if isinstance(fig_dict, dict) else type(fig_dict)}")
            b = io.BytesIO(); err_img.save(b, format='PNG'); b.seek(0)
            return dict(content=base64.b64encode(b.read()).decode(), filename='project_tree_error.png', type='image/png')
        fig = go.Figure(fig_dict)
        # Try high-quality export using kaleido (explicit engine) first
        base_png = None
        try:
            base_png = fig.to_image(format='png', scale=2, engine='kaleido')
        except Exception as err:
            try:
                import kaleido  # noqa: F401
                kaleido_present = True
            except Exception:
                kaleido_present = False
            print(f"[png][diag][error] to_image failed err={err} kaleido_present={kaleido_present}")
            # Likely kaleido missing; create fallback image
            fallback = PILImage.new('RGB',(1200,400),'white')
            d = ImageDraw.Draw(fallback)
            msg = 'Static export requires kaleido. Install with: pip install kaleido'
            try:
                font = ImageFont.truetype('Arial.ttf',16)
            except Exception:
                font = ImageFont.load_default()
            d.text((20,20), msg, fill='black', font=font)
            buff_fb = io.BytesIO()
            fallback.save(buff_fb, format='PNG')
            buff_fb.seek(0)
            png_fallback_bytes = buff_fb.read()
            return dcc.send_bytes(lambda b: b.write(png_fallback_bytes), 'project_tree_missing_kaleido.png')
        # Defensive: ensure we received bytes and PNG signature
        if not base_png or not isinstance(base_png, (bytes, bytearray)):
            raise ValueError('Unexpected base_png type from fig.to_image')
        if not base_png.startswith(b'\x89PNG\r\n\x1a\n'):
            print('[png][diag][warn] Base image missing PNG signature – attempting Pillow open & re-save')
        base_img = PILImage.open(io.BytesIO(base_png)).convert('RGBA')
        width = base_img.width
        header_h = 110
        footer_h = 40
        total_h = header_h + base_img.height + footer_h
        out_img = PILImage.new('RGBA', (width, total_h), (255,255,255,255))
        draw = ImageDraw.Draw(out_img)
        logo_path = Path('assets') / 'LSI_Logo.svg'
        try:
            import cairosvg, tempfile
            if logo_path.exists():
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                    cairosvg.svg2png(url=str(logo_path), write_to=tmp.name, output_width=360)
                    logo_png = PILImage.open(tmp.name).convert('RGBA')
                    lh = 95
                    ratio = lh / logo_png.height
                    lw = int(logo_png.width * ratio)
                    logo_png = logo_png.resize((lw, lh))
                    # Center horizontally
                    out_img.paste(logo_png, ((width - lw)//2, int((header_h - lh)/2)), logo_png)
        except Exception:
            # Fallback simple text if svg -> png conversion unavailable
            try:
                from PIL import ImageFont
                fnt = ImageFont.load_default()
                draw.text((width//2 - 40, 30), 'LSI', fill=(40,40,40,255), font=fnt)
            except Exception:
                pass
        title_text = 'Project Tree'
        try:
            font = ImageFont.truetype('Arial.ttf', 26)
        except Exception:
            font = ImageFont.load_default()
        try:
            bbox = draw.textbbox((0,0), title_text, font=font)
            tw = bbox[2]-bbox[0]; th = bbox[3]-bbox[1]
        except Exception:
            tw, th = draw.textlength(title_text, font=font), 26
        draw.text(((width - tw)//2, (header_h - th)//2), title_text, fill=(20,20,20,255), font=font)
        out_img.paste(base_img, (0, header_h))
        footer_y = header_h + base_img.height
        draw.line([(0, footer_y),(width, footer_y)], fill=(180,180,180,255), width=1)
        footer_text = '© 2025 LSI Graphics, LLC — Generated by Sign Package Estimator'
        try:
            ffont = ImageFont.truetype('Arial.ttf', 14)
        except Exception:
            ffont = ImageFont.load_default()
        try:
            fbbox = draw.textbbox((0,0), footer_text, font=ffont)
            ftw = fbbox[2]-fbbox[0]; fth = fbbox[3]-fbbox[1]
        except Exception:
            ftw = draw.textlength(footer_text, font=ffont); fth = 14
        draw.text((width - ftw - 20, footer_y + (footer_h - fth)//2), footer_text, fill=(80,80,80,255), font=ffont)
        buff = io.BytesIO()
        out_img.convert('RGB').save(buff, format='PNG')
        buff.seek(0)
        png_bytes = buff.read()
        sig = png_bytes[:8]
        valid_sig = sig == b'\x89PNG\r\n\x1a\n'
        sha10 = __import__('hashlib').sha1(png_bytes).hexdigest()[:10]
        print(f"[png][diag] bytes={len(png_bytes)} sig={sig!r} valid_sig={valid_sig} sha1={sha10}")
        if not valid_sig:
            # Attempt force re-encode to fix
            try:
                fix_img = PILImage.open(io.BytesIO(png_bytes)).convert('RGB')
                rebuff = io.BytesIO(); fix_img.save(rebuff, format='PNG'); rebuff.seek(0)
                png_bytes = rebuff.read(); sig = png_bytes[:8]; valid_sig = sig == b'\x89PNG\r\n\x1a\n'
                print(f"[png][diag] re-encoded valid_sig={valid_sig}")
            except Exception as _re:
                print(f"[png][diag][warn] re-encode failed: {_re}")
        return dcc.send_bytes(lambda b: b.write(png_bytes), 'project_tree.png')
    except Exception as e:
        print(f"[export-tree-png][error] {e}")
        raise PreventUpdate

# ------------------ Tree View Mode Switch (Cytoscape) ------------------ #
@app.callback(
    Output('project-tree-wrapper','children'),
    Input('tree-view-mode','value'),
    State('cyto-spacing-factor','value'),
    State('cyto-padding','value'),
    prevent_initial_call=True
)
def switch_tree_view(mode, spacing_factor, padding):
    # Guard against pandas Series / unexpected iterable causing ambiguous truth errors
    try:
        if isinstance(mode, pd.Series):
            # Take first value if series provided
            mode = mode.iloc[0] if not mode.empty else 'static'
    except Exception:
        pass
    if mode == 'cyto':
        nodes = get_project_tree_data()
        # Preload sign types metadata
        conn = sqlite3.connect(DATABASE_PATH)
        try:
            st_df = pd.read_sql_query('SELECT name, unit_price, width, height, price_per_sq_ft, material_multiplier, material, description, image_path FROM sign_types', conn)
        except Exception:
            st_df = pd.DataFrame(columns=['name','unit_price','width','height','price_per_sq_ft','material_multiplier','material','description'])
        finally:
            conn.close()
        st_map = {r['name']: r for _, r in st_df.iterrows()}
        elements = []
        id_map = {n['id']: n for n in nodes}
        for n in nodes:
            data = {'id': n['id'], 'label': n['label'], 'type': n['type']}
            if n['type'] == 'sign':
                raw = n['label']
                base_name = raw.split('(')[0].strip()
                if base_name.startswith('Group:'):
                    base_name = base_name.replace('Group:','').strip()
                info = st_map.get(base_name)
                # Explicit None check; avoid ambiguous Series truth evaluation
                if info is not None:
                    width = info.get('width') or 0
                    height = info.get('height') or 0
                    area = (width or 0) * (height or 0)
                    data.update({
                        'sign_name': base_name,
                        'material': info.get('material') or '',
                        'description': (info.get('description') or '')[:160],
                        'width': width,
                        'height': height,
                        'area': area,
                        'unit_price': info.get('unit_price') or 0,
                        'price_per_sq_ft': info.get('price_per_sq_ft') or 0,
                        'material_multiplier': info.get('material_multiplier') or 0,
                        'image_path': info.get('image_path') or ''
                    })
            elements.append({'data': data, 'classes': n['type']})
        for n in nodes:
            parent = n.get('parent')
            if parent and parent in id_map:
                elements.append({'data': {'source': parent, 'target': n['id']}})
        stylesheet = [
            {'selector': 'node','style': {'content':'data(label)','text-wrap':'wrap','text-max-width':120,'text-valign':'center','color':'#fff','font-size':'9px','background-color':'#4a90e2','padding':'4px'}},
            {'selector': '.project','style': {'background-color':'#1f77b4','font-size':'11px','font-weight':'bold'}},
            {'selector': '.building','style': {'background-color':'#ff7f0e'}},
            {'selector': '.sign','style': {'background-color':'#2ca02c'}},
            {'selector': 'edge','style': {'width':2,'line-color':'#ccc','curve-style':'bezier'}}
        ]
        # Fallback defaults if None
        try:
            sf = float(spacing_factor) if spacing_factor else 1.25
        except Exception:
            sf = 1.25
        try:
            pad = int(padding) if padding is not None else 15
        except Exception:
            pad = 15
        cyto_component = cyto.Cytoscape(
            id='project-tree-cyto',
            elements=elements,
            layout={'name':'breadthfirst','directed':True,'spacingFactor':sf,'padding':pad},
            style={'width':'100%','height':'600px'},
            stylesheet=stylesheet,
            generateImage=False,
            zoom=1,
            minZoom=0.2,
            maxZoom=4
        )
        return [
            cyto_component,
            html.Div(id='cyto-hover-panel', style={'position':'absolute','top':'60px','right':'25px','zIndex':1050,'display':'none','maxWidth':'340px'})
        ]
    return dcc.Graph(id='project-tree', figure=safe_tree_figure())

# Initial tree population when app loads / Projects tab first shown
@app.callback(
    Output('project-tree','figure'),
    Input('main-tabs','active_tab')
)
def init_tree(active_tab):
    if active_tab == 'projects-tab':
        return safe_tree_figure()
    raise PreventUpdate

# Show/hide cytoscape layout controls based on selected mode
@app.callback(
    Output('cyto-layout-controls','style'),
    Input('tree-view-mode','value')
)
def toggle_layout_controls(mode):
    if mode == 'cyto':
        return {'display':'block'}
    return {'display':'none'}

# Update Cytoscape layout dynamically when spacing sliders move
@app.callback(
    Output('project-tree-cyto','layout'),
    Output('project-tree-cyto','stylesheet', allow_duplicate=True),
    Input('cyto-spacing-factor','value'),
    Input('cyto-padding','value'),
    Input('cyto-label-width','value'),
    Input('cyto-node-size','value'),
    Input('cyto-toggle-labels','value'),
    Input('cyto-toggle-images','value'),
    State('project-tree-cyto','stylesheet'),
    State('tree-view-mode','value'),
    prevent_initial_call=True
)
def update_cyto_layout(spacing_factor, padding, label_width, node_size, toggle_labels, toggle_images, current_stylesheet, mode):
    if mode != 'cyto':
        raise PreventUpdate
    try:
        sf = float(spacing_factor) if spacing_factor else 1.25
    except Exception:
        sf = 1.25
    try:
        pad = int(padding) if padding is not None else 15
    except Exception:
        pad = 15
    try:
        lw = int(label_width) if label_width else 120
    except Exception:
        lw = 120
    try:
        ns = int(node_size) if node_size else 15
    except Exception:
        ns = 15
    show_labels = bool(toggle_labels and 'labels' in toggle_labels)
    show_images = bool(toggle_images and 'images' in toggle_images)
    base_styles = [
        {'selector':'node','style':{
            'content': 'data(label)' if show_labels else '',
            'text-wrap':'wrap','text-max-width': lw,
            'text-valign':'center','color':'#fff','font-size':'9px',
            'background-color':'#4a90e2','padding':'4px',
            'width': ns, 'height': ns,
            'background-fit': 'cover'
        }},
        {'selector':'.project','style':{'background-color':'#1f77b4','font-size':'11px','font-weight':'bold','width':ns+6,'height':ns+6}},
        {'selector':'.building','style':{'background-color':'#ff7f0e'}},
        {'selector':'.sign','style':{'background-color':'#2ca02c'}},
        {'selector':'edge','style':{'width':2,'line-color':'#bbb','curve-style':'bezier'}},
        {'selector':'node.hover','style':{'border-width':3,'border-color':'#222','shadow-blur':8,'shadow-color':'#999','shadow-opacity':0.7,'shadow-offset-x':2,'shadow-offset-y':2}}
    ]
    if show_images:
        # use image_path as background if present
        base_styles.append({'selector':'node[image_path]','style':{'background-image':'data(image_path)','background-color':'#1f1f1f'}})
    return {'name':'breadthfirst','directed':True,'spacingFactor':sf,'padding':pad}, base_styles

## (Removed earlier duplicate hover callback to avoid duplicate output errors)

### Diagnostics banner removed per request (SVG Export Notice & Optional Packages Missing)

# Project deletion callback
@app.callback(
    Output('projects-list','children', allow_duplicate=True),
    Output('project-create-feedback','children', allow_duplicate=True),
    Output('project-tree','figure', allow_duplicate=True),
    Output('assign-project-dropdown','options', allow_duplicate=True),
    Output('project-edit-dropdown','options', allow_duplicate=True),
    Output('projects-debug-list','children', allow_duplicate=True),
    Input('delete-project-confirm','submit_n_clicks'),
    State('project-edit-dropdown','value'),
    prevent_initial_call=True
)
def delete_project(confirm_clicks, project_id):
    if not confirm_clicks:
        raise PreventUpdate
    if not project_id:
        return (dash.no_update,
                dbc.Alert('Select a project to delete', color='warning'),
                dash.no_update,
                dash.no_update,
                dash.no_update,
                dash.no_update)
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        print(f"[delete_project] Deleting project id={project_id}")
        cur.execute('DELETE FROM building_sign_groups WHERE building_id IN (SELECT id FROM buildings WHERE project_id=?)', (project_id,))
        cur.execute('DELETE FROM building_signs WHERE building_id IN (SELECT id FROM buildings WHERE project_id=?)', (project_id,))
        cur.execute('DELETE FROM buildings WHERE project_id=?', (project_id,))
        cur.execute('DELETE FROM projects WHERE id=?', (project_id,))
        conn.commit()
        df = pd.read_sql_query('SELECT id, name, created_date FROM projects ORDER BY id DESC', conn)
        conn.close()
        if df.empty:
            list_children = html.Div('No projects yet.')
            options = []
        else:
            rows = [html.Li(f"{r.name} (ID {r.id}) - {r.created_date}") for r in df.itertuples()]
            list_children = html.Ul(rows, className='mb-0')
            options = [{'label': r.name, 'value': r.id} for r in df.itertuples()]
        debug_txt = ' | '.join(f"{r.id}:{r.name}" for r in df.itertuples()) if not df.empty else 'none'
        return (list_children,
                dbc.Alert('Project deleted', color='info'),
                safe_tree_figure(),
                options,
                options,
                debug_txt)
    except Exception as e:
        return (dash.no_update,
                dbc.Alert(f'Error deleting: {e}', color='danger'),
                dash.no_update,
                dash.no_update,
                dash.no_update,
                dash.no_update)

# Show delete confirmation dialog
@app.callback(
    Output('delete-project-confirm','displayed'),
    Output('project-create-feedback','children', allow_duplicate=True),
    Input('delete-project-btn','n_clicks'),
    State('project-edit-dropdown','value'),
    prevent_initial_call=True
)
def show_delete_confirm(n_clicks, project_id):
    if not n_clicks:
        raise PreventUpdate
    if not project_id:
        return False, dbc.Alert('Select a project to delete first', color='warning')
    return True, dash.no_update

# Dynamic disabling of install parameter inputs + hint text
@app.callback(
    Output('install-percent-input','disabled'),
    Output('install-per-sign-rate','disabled'),
    Output('install-per-area-rate','disabled'),
    Output('install-hours-input','disabled'),
    Output('install-hourly-rate-input','disabled'),
    Output('install-mode-hint','children'),
    Input('install-mode','value')
)
def update_install_inputs(mode):
    if not mode:
        raise PreventUpdate
    percent_dis = mode != 'percent'
    per_sign_dis = mode != 'per_sign'
    per_area_dis = mode != 'per_area'
    hours_hours_dis = mode != 'hours'
    hours_rate_dis = mode != 'hours'
    hints = {
        'percent': 'Percent: Uses % of subtotal for installation.',
        'per_sign': 'Per Sign: Uses Install $/Sign * total signs (auto if enabled when blank).',
        'per_area': 'Per Area: Install $/SqFt * total sign area.',
        'hours': 'Hours: (Hours or auto sum) * $/Hour.',
        'none': 'None: No installation cost added.'
    }
    return percent_dis, per_sign_dis, per_area_dis, hours_hours_dis, hours_rate_dis, hints.get(mode,'')

# ------------------ Sign Types Table CRUD ------------------ #
@app.callback(
    Output('signs-table', 'data', allow_duplicate=True),
    Output('signs-save-status', 'children'),
    Output('signs-table-master-store','data', allow_duplicate=True),
    Input('main-tabs', 'active_tab'),
    Input('signs-table', 'data_timestamp'),
    Input('add-sign-btn', 'n_clicks'),
    State('signs-table', 'data'),
    prevent_initial_call='initial_duplicate'
)
def manage_sign_types(active_tab, data_ts, add_clicks, data_rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if 'main-tabs' in triggered and active_tab == 'signs-tab':
        conn = sqlite3.connect(DATABASE_PATH)
        df = pd.read_sql_query(
            "SELECT name, description, material_alt, unit_price, material, price_per_sq_ft, material_multiplier, width, height, install_type, install_time_hours, per_sign_install_rate, image_path FROM sign_types ORDER BY name",
            conn
        )
        conn.close()
    records = df.to_dict('records')
    return records, '', records
    if 'add-sign-btn' in triggered:
        rows = data_rows or []
        rows.append({"name":"","description":"","unit_price":0,"material":"","price_per_sq_ft":0,"width":0,"height":0})
    return rows, 'New row added', rows
    if 'signs-table' in triggered and active_tab == 'signs-tab':
        rows = data_rows or []
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        saved = 0
        cleaned = []
        for row in rows:
            name = (row.get('name') or '').strip()
            if not name:
                continue
            def n(v):
                try: return float(v or 0)
                except: return 0.0
            # New extended columns (may not yet be in table for older deployments; ignore errors silently)
            extended_sql = '''
                INSERT INTO sign_types (name, description, material_alt, unit_price, material, price_per_sq_ft, width, height, material_multiplier, install_type, install_time_hours, per_sign_install_rate, image_path)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(name) DO UPDATE SET description=excluded.description, material_alt=excluded.material_alt, unit_price=excluded.unit_price,
                    material=excluded.material, price_per_sq_ft=excluded.price_per_sq_ft, width=excluded.width, height=excluded.height,
                    material_multiplier=excluded.material_multiplier, install_type=excluded.install_type, install_time_hours=excluded.install_time_hours,
                    per_sign_install_rate=excluded.per_sign_install_rate, image_path=COALESCE(excluded.image_path, image_path)
            '''
            material_alt = (row.get('material_alt') or row.get('Material2') or '')[:120]
            material_multiplier = n(row.get('material_multiplier'))
            install_type_val = (row.get('install_type') or '')[:60]
            install_time_hours = n(row.get('install_time_hours') or row.get('install_time'))
            per_sign_install_rate = n(row.get('per_sign_install_rate'))
            try:
                cur.execute(extended_sql, (
                    name,
                    (row.get('description') or '')[:255],
                    material_alt,
                    n(row.get('unit_price')),
                    (row.get('material') or '')[:120],
                    n(row.get('price_per_sq_ft')),
                    n(row.get('width')),
                    n(row.get('height')),
                    material_multiplier,
                    install_type_val,
                    install_time_hours,
                    per_sign_install_rate,
                    (row.get('image_path') or None)
                ))
            except Exception:
                # Fallback to legacy column set if migration not applied
                cur.execute('''
                    INSERT INTO sign_types (name, description, unit_price, material, price_per_sq_ft, width, height)
                    VALUES (?,?,?,?,?,?,?)
                    ON CONFLICT(name) DO UPDATE SET description=excluded.description, unit_price=excluded.unit_price,
                        material=excluded.material, price_per_sq_ft=excluded.price_per_sq_ft, width=excluded.width, height=excluded.height
                ''', (
                    name,
                    (row.get('description') or '')[:255],
                    n(row.get('unit_price')),
                    (row.get('material') or '')[:120],
                    n(row.get('price_per_sq_ft')),
                    n(row.get('width')),
                    n(row.get('height'))
                ))
            saved += 1
            cleaned.append(row)
        conn.commit(); conn.close()
        return cleaned, f'Saved {saved} rows', cleaned
    raise PreventUpdate

# Dynamic page size control for sign types table
@app.callback(
    Output('signs-table','page_size'),
    Input('signs-page-size-dropdown','value'),
    State('signs-table','data'),
    prevent_initial_call=True
)
def update_signs_page_size(size, rows):
    if size is None:
        raise PreventUpdate
    if size == -1:
        # All rows
        return len(rows) if rows else 10000
    return size

# Filter sign types by install type (ext vs non-ext)
@app.callback(
    Output('signs-table','data', allow_duplicate=True),
    Input('signs-install-filter','value'),
    Input('signs-table-master-store','data'),
    prevent_initial_call='initial_duplicate'
)
def filter_signs_by_install(filter_value, master_rows):
    if filter_value is None or filter_value == 'all':
        return master_rows
    try:
        if filter_value == 'ext':
            return [r for r in master_rows if (r.get('install_type') or '').strip().lower() == 'ext']
        elif filter_value == 'non_ext':
            return [r for r in master_rows if (r.get('install_type') or '').strip().lower() != 'ext']
        return master_rows
    except Exception as e:
        print(f"[filter_signs_by_install][error] {e}")
        return master_rows

## Legacy single-image upload callback removed in favor of multi-image system above.

# Serve uploaded sign images (simple static route)
try:
    from flask import send_file, abort
    @app.server.route('/sign-images/<path:filename>')
    def serve_sign_image(filename):
        base_dir = Path('sign_images')
        target = (base_dir / filename).resolve()
        # Prevent path traversal: ensure parent
        if not str(target).startswith(str(base_dir.resolve())):
            return abort(404)
        if not target.exists():
            return abort(404)
        # Infer mimetype
        ext = target.suffix.lower().lstrip('.')
        mime = {
            'png':'image/png',
            'jpg':'image/jpeg',
            'jpeg':'image/jpeg',
            'gif':'image/gif',
            'svg':'image/svg+xml'
        }.get(ext, 'application/octet-stream')
        return send_file(str(target), mimetype=mime)
except Exception as e:
    print(f"[sign-image][route][warn] Could not register image route: {e}")

# Tiny debug stats updater (tab focused or after save/import) 
@app.callback(
    Output('debug-signs-stats','children'),
    Input('main-tabs','active_tab'),
    Input('signs-save-status','children'),
    Input('material-pricing-feedback','children'),
    prevent_initial_call=True
)
def update_debug_stats(active_tab, sign_save_msg, material_msg):
    if active_tab != 'signs-tab':
        raise PreventUpdate
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        cur.execute('SELECT COUNT(*) FROM sign_types')
        sign_count = cur.fetchone()[0]
        cur.execute('SELECT COUNT(*) FROM material_pricing')
        mat_count = cur.fetchone()[0]
        cur.execute("SELECT name FROM sign_types ORDER BY last_modified DESC LIMIT 3")
        recent_signs = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT material_name FROM material_pricing ORDER BY last_updated DESC LIMIT 3")
        recent_mats = [r[0] for r in cur.fetchall()]
        conn.close()
        return f"sign_types: {sign_count} (recent: {', '.join(recent_signs) if recent_signs else 'n/a'}) | materials: {mat_count} (recent: {', '.join(recent_mats) if recent_mats else 'n/a'})"
    except Exception as e:
        return f"debug error: {e}"

# Allow closing the debug card
@app.callback(
    Output('debug-signs-card','style'),
    Input('close-debug-stats','n_clicks'),
    prevent_initial_call=True
)
def hide_debug_card(n):
    if not n:
        raise PreventUpdate
    return {'display':'none'}

# ------------------ Cytoscape Hover Panel ------------------ #
@app.callback(
    Output('cyto-hover-panel','children'),
    Output('cyto-hover-panel','style'),
    Input('project-tree-wrapper','children'),  # placeholder to allow dynamic component existence
    Input('project-tree-cyto','mouseoverData'),
    prevent_initial_call=True
)
def cyto_hover(_wrapper_children, mouseover):
    # If cytoscape not rendered yet or no hover data
    if not mouseover or 'data' not in mouseover:
        raise PreventUpdate
    data = mouseover.get('data') or {}
    if data.get('type') != 'sign':
        # Hide panel for non-sign nodes
        return dash.no_update, {'display':'none'}
    # Build panel
    def fmt(num):
        try:
            return f"{float(num):,.2f}"
        except Exception:
            return str(num)
    header = data.get('sign_name') or data.get('label')
    # Build image (if any)
    img_el = None
    img_rel = data.get('image_path')
    if img_rel:
        # Normalize path for route (strip leading ./ or backslashes)
        p = str(img_rel).replace('\\','/')
        if p.startswith('./'):
            p = p[2:]
        if p.startswith('sign_images/'):
            fname = p.split('/',1)[1]
        else:
            # assume stored as sign_images/filename or raw filename
            fname = Path(p).name
        img_el = html.Img(src=f"/sign-images/{fname}", style={'maxWidth':'100%','maxHeight':'140px','objectFit':'contain','marginBottom':'8px','border':'1px solid #ddd','padding':'2px','background':'#fff'})
    body = dbc.Card([
        dbc.CardHeader(html.Strong(header)),
        dbc.CardBody([
            img_el,
            html.Div(data.get('description') or '', className='mb-2 small'),
            html.Table([
                html.Tbody([
                    html.Tr([html.Th('Material', className='pe-2'), html.Td(data.get('material') or '-')]),
                    html.Tr([html.Th('Width'), html.Td(fmt(data.get('width')))]),
                    html.Tr([html.Th('Height'), html.Td(fmt(data.get('height')))]),
                    html.Tr([html.Th('Area'), html.Td(fmt(data.get('area')))]),
                    html.Tr([html.Th('Unit Price'), html.Td(f"$ {fmt(data.get('unit_price'))}")]),
                    html.Tr([html.Th('$ / SqFt'), html.Td(fmt(data.get('price_per_sq_ft')))]),
                    html.Tr([html.Th('Multiplier'), html.Td(fmt(data.get('material_multiplier')))])
                ])
            ], className='table table-sm mb-0')
        ])
    ], className='shadow-sm')
    style = {'position':'absolute','top':'60px','right':'25px','zIndex':1050,'display':'block','maxWidth':'340px'}
    return body, style

# ------------------ Static Plotly Tree Hover Panel ------------------ #
@app.callback(
    Output('cyto-hover-panel','children', allow_duplicate=True),
    Output('cyto-hover-panel','style', allow_duplicate=True),
    Input('project-tree','hoverData'),
    prevent_initial_call=True
)
def static_tree_hover(hoverData):
    if not hoverData or 'points' not in hoverData:
        raise PreventUpdate
    pt = (hoverData['points'] or [{}])[0]
    label = pt.get('text') or pt.get('customdata') or ''
    # Derive base sign name (strip qty)
    base = str(label).split('(')[0].strip()
    if not base:
        raise PreventUpdate
    # Fetch sign details including image
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        sdf = pd.read_sql_query('SELECT name, description, material, width, height, unit_price, price_per_sq_ft, material_multiplier, image_path FROM sign_types WHERE LOWER(name)=LOWER(?)', conn, params=(base,))
        conn.close()
    except Exception as e:
        print(f"[static-hover][error] {e}")
        raise PreventUpdate
    if sdf.empty:
        raise PreventUpdate
    r = sdf.iloc[0]
    def fmt(num):
        try:
            return f"{float(num):,.2f}"
        except Exception:
            return str(num)
    area = (r.get('width') or 0) * (r.get('height') or 0)
    img_el=None
    img_rel = r.get('image_path')
    if img_rel:
        p = str(img_rel).replace('\\','/')
        fname = Path(p).name
        img_el = html.Img(src=f"/sign-images/{fname}", style={'maxWidth':'100%','maxHeight':'140px','objectFit':'contain','marginBottom':'8px','border':'1px solid #ddd','padding':'2px','background':'#fff'})
    body = dbc.Card([
        dbc.CardHeader(html.Strong(base)),
        dbc.CardBody([
            img_el,
            html.Div((r.get('description') or '')[:160], className='mb-2 small'),
            html.Table([
                html.Tbody([
                    html.Tr([html.Th('Material', className='pe-2'), html.Td(r.get('material') or '-')]),
                    html.Tr([html.Th('Width'), html.Td(fmt(r.get('width')))]),
                    html.Tr([html.Th('Height'), html.Td(fmt(r.get('height')))]),
                    html.Tr([html.Th('Area'), html.Td(fmt(area))]),
                    html.Tr([html.Th('Unit Price'), html.Td(f"$ {fmt(r.get('unit_price'))}")]),
                    html.Tr([html.Th('$ / SqFt'), html.Td(fmt(r.get('price_per_sq_ft')))]),
                    html.Tr([html.Th('Multiplier'), html.Td(fmt(r.get('material_multiplier')))])
                ])
            ], className='table table-sm mb-0')
        ])
    ], className='shadow-sm')
    style = {'position':'absolute','top':'60px','right':'25px','zIndex':1050,'display':'block','maxWidth':'340px'}
    return body, style

# ------------------ Material Pricing CRUD & Recalc ------------------ #
@app.callback(
    Output('material-pricing-table','data'),
    Output('material-pricing-feedback','children'),
    Input('main-tabs','active_tab'),
    Input('add-material-btn','n_clicks'),
    Input('save-materials-btn','n_clicks'),
    Input('recalc-sign-prices-btn','n_clicks'),
    State('material-pricing-table','data'),
    prevent_initial_call=True
)
def manage_material_pricing(active_tab, add_clicks, save_clicks, recalc_clicks, rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if 'main-tabs' in triggered and active_tab == 'signs-tab':
        conn = sqlite3.connect(DATABASE_PATH)
        try:
            df = pd.read_sql_query("SELECT material_name, price_per_sq_ft FROM material_pricing ORDER BY material_name", conn)
        except Exception as e:
            conn.close()
            return [], dbc.Alert(f"Error loading materials: {e}", color='danger')
        conn.close()
        return df.to_dict('records'), ''
    if 'add-material-btn' in triggered:
        data = rows or []
        data.append({'material_name':'','price_per_sq_ft':0})
        return data, 'Row added'
    if 'save-materials-btn' in triggered and rows is not None:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        saved = 0
        for r in rows:
            name = (r.get('material_name') or '').strip()
            if not name:
                continue
            try:
                p = float(r.get('price_per_sq_ft') or 0)
            except:
                p = 0
            cur.execute('''
                INSERT INTO material_pricing (material_name, price_per_sq_ft)
                VALUES (?,?)
                ON CONFLICT(material_name) DO UPDATE SET price_per_sq_ft=excluded.price_per_sq_ft, last_updated=CURRENT_TIMESTAMP
            ''', (name, p))
            saved += 1
        conn.commit(); conn.close()
        return rows, f'Saved {saved} materials'
    if 'recalc-sign-prices-btn' in triggered:
        conn = sqlite3.connect(DATABASE_PATH)
        cur = conn.cursor()
        cur.execute('''
            UPDATE sign_types
            SET unit_price = CASE 
                WHEN width>0 AND height>0 THEN (
                    width*height*COALESCE(
                        (SELECT price_per_sq_ft FROM material_pricing mp WHERE LOWER(mp.material_name)=LOWER(sign_types.material)),
                        price_per_sq_ft
                    )
                ) 
                ELSE unit_price END,
                price_per_sq_ft = COALESCE((SELECT price_per_sq_ft FROM material_pricing mp WHERE LOWER(mp.material_name)=LOWER(sign_types.material)), price_per_sq_ft),
                last_modified = CURRENT_TIMESTAMP
        ''')
        conn.commit(); conn.close()
        return rows, 'Recalculated sign prices'
    raise PreventUpdate

# ------------------ Sign Groups CRUD ------------------ #
@app.callback(
    Output('group-save-feedback','children'),
    Output('group-select-dropdown','options', allow_duplicate=True),
    Output('group-assign-group-dropdown','options', allow_duplicate=True),
    Input('save-group-btn','n_clicks'),
    State('group-name-input','value'),
    State('group-desc-input','value'),
    prevent_initial_call=True
)
def save_group(n_clicks, name, desc):
    if not n_clicks:
        raise PreventUpdate
    name = (name or '').strip()
    if not name:
        return dbc.Alert('Name required', color='danger'), dash.no_update, dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    try:
        cur.execute('''
            INSERT INTO sign_groups (name, description) VALUES (?,?)
            ON CONFLICT(name) DO UPDATE SET description=excluded.description
        ''', (name, (desc or '')[:255]))
        conn.commit()
        groups_df = pd.read_sql_query('SELECT id, name FROM sign_groups ORDER BY name', conn)
        conn.close()
        options = [{'label': r.name, 'value': r.id} for r in groups_df.itertuples()]
        return dbc.Alert(f"Group '{name}' saved", color='success'), options, options
    except Exception as e:
        conn.close()
        return dbc.Alert(f'Error: {e}', color='danger'), dash.no_update, dash.no_update

@app.callback(
    Output('group-members-table','data'),
    Output('group-members-feedback','children', allow_duplicate=True),
    Input('group-select-dropdown','value'),
    Input('group-add-sign-btn','n_clicks'),
    Input('group-save-members-btn','n_clicks'),
    State('group-add-sign-dropdown','value'),
    State('group-add-sign-qty','value'),
    State('group-members-table','data'),
    prevent_initial_call=True
)
def manage_group_members(group_id, add_clicks, save_clicks, sign_type_id, qty, rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if not group_id:
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    feedback = dash.no_update
    if 'group-add-sign-btn' in triggered and sign_type_id:
        cur.execute('SELECT id FROM sign_group_members WHERE group_id=? AND sign_type_id=?', (group_id, sign_type_id))
        ex = cur.fetchone()
        q = max(1, int(qty or 1))
        if ex:
            cur.execute('UPDATE sign_group_members SET quantity=? WHERE id=?', (q, ex[0]))
        else:
            cur.execute('INSERT INTO sign_group_members (group_id, sign_type_id, quantity) VALUES (?,?,?)', (group_id, sign_type_id, q))
        conn.commit()
        feedback = 'Member added/updated'
    elif 'group-save-members-btn' in triggered and rows:
        for r in rows:
            name = r.get('sign_name')
            q = max(0, int(r.get('quantity') or 0))
            cur.execute('SELECT id FROM sign_types WHERE name=?', (name,))
            st = cur.fetchone()
            if not st:
                continue
            cur.execute('SELECT id FROM sign_group_members WHERE group_id=? AND sign_type_id=?', (group_id, st[0]))
            ex = cur.fetchone()
            if ex:
                cur.execute('UPDATE sign_group_members SET quantity=? WHERE id=?', (q, ex[0]))
        conn.commit()
        feedback = 'Member quantities saved'
    # Load
    df = pd.read_sql_query('''SELECT st.name as sign_name, sgm.quantity FROM sign_group_members sgm JOIN sign_types st ON sgm.sign_type_id=st.id WHERE sgm.group_id=? ORDER BY st.name''', conn, params=(group_id,))
    conn.close()
    return df.to_dict('records'), feedback

# ------------------ Assign Groups to Buildings ------------------ #
@app.callback(
    Output('group-assign-project-dropdown','options'),
    Input('main-tabs','active_tab')
)
def populate_group_project_options(active_tab):
    if active_tab != 'groups-tab':
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('SELECT id, name FROM projects ORDER BY name', conn)
    conn.close()
    return [{'label': r.name, 'value': r.id} for r in df.itertuples()]

@app.callback(
    Output('group-assign-building-dropdown','options'),
    Output('group-assign-building-dropdown','value'),
    Input('group-assign-project-dropdown','value')
)
def populate_group_buildings(project_id):
    if not project_id:
        return [], None
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('SELECT id, name FROM buildings WHERE project_id=? ORDER BY name', conn, params=(project_id,))
    conn.close()
    opts = [{'label': r.name,'value': r.id} for r in df.itertuples()]
    return opts, (opts[0]['value'] if opts else None)

@app.callback(
    Output('building-groups-table','data'),
    Output('group-assign-feedback','children'),
    Input('group-assign-building-dropdown','value'),
    Input('group-assign-btn','n_clicks'),
    Input('building-save-group-qty-btn','n_clicks'),
    State('group-assign-group-dropdown','value'),
    State('group-assign-qty','value'),
    State('building-groups-table','data'),
    prevent_initial_call=True
)
def manage_building_groups(building_id, add_clicks, save_clicks, group_id, qty, rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if not building_id:
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    feedback = dash.no_update
    if 'group-assign-btn' in triggered and group_id:
        cur.execute('SELECT id FROM building_sign_groups WHERE building_id=? AND group_id=?', (building_id, group_id))
        ex = cur.fetchone()
        q = max(1, int(qty or 1))
        if ex:
            cur.execute('UPDATE building_sign_groups SET quantity=? WHERE id=?', (q, ex[0]))
        else:
            cur.execute('INSERT INTO building_sign_groups (building_id, group_id, quantity) VALUES (?,?,?)', (building_id, group_id, q))
        conn.commit()
        feedback = 'Group assigned'
    elif 'building-save-group-qty-btn' in triggered and rows:
        for r in rows:
            name = r.get('group_name')
            q = max(0, int(r.get('quantity') or 0))
            cur.execute('SELECT id FROM sign_groups WHERE name=?', (name,))
            gr = cur.fetchone()
            if not gr: continue
            cur.execute('SELECT id FROM building_sign_groups WHERE building_id=? AND group_id=?', (building_id, gr[0]))
            ex = cur.fetchone()
            if ex:
                cur.execute('UPDATE building_sign_groups SET quantity=? WHERE id=?', (q, ex[0]))
        conn.commit(); feedback='Group quantities saved'
    df = pd.read_sql_query('''SELECT sg.name as group_name, bsg.quantity FROM building_sign_groups bsg JOIN sign_groups sg ON bsg.group_id=sg.id WHERE bsg.building_id=? ORDER BY sg.name''', conn, params=(building_id,))
    conn.close()
    return df.to_dict('records'), feedback

# ------------------ Building View Tab Callbacks ------------------ #
@app.callback(
    Output('bv-building-dropdown','options'),
    Output('bv-building-dropdown','value'),
    Input('bv-project-dropdown','value'),
    prevent_initial_call=True
)
def bv_load_buildings(project_id):
    if not project_id:
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('SELECT id, name FROM buildings WHERE project_id=? ORDER BY name', conn, params=(project_id,))
    conn.close()
    opts = [{'label': r.name, 'value': r.id} for r in df.itertuples()]
    return opts, (opts[0]['value'] if opts else None)

@app.callback(
    Output('bv-sign-type-dropdown','options'),
    Output('bv-signs-table','data'),
    Output('bv-delete-sign-dropdown','options'),
    Output('bv-delete-group-dropdown','options'),
    Output('bv-building-meta','children'),
    Output('bv-summary','children'),
    Input('bv-building-dropdown','value'),
    prevent_initial_call=True
)
def bv_load_building(building_id):
    if not building_id:
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    st_df = pd.read_sql_query('SELECT id, name, unit_price FROM sign_types ORDER BY name', conn)
    b_df = pd.read_sql_query('SELECT name, description FROM buildings WHERE id=?', conn, params=(building_id,))
    rows_df = pd.read_sql_query('''SELECT st.name as sign_name, bs.quantity, st.unit_price, (bs.quantity*st.unit_price) as total
                                   FROM building_signs bs JOIN sign_types st ON bs.sign_type_id=st.id
                                   WHERE bs.building_id=? ORDER BY st.name''', conn, params=(building_id,))
    grp_df = pd.read_sql_query('''SELECT sg.name, bsg.quantity FROM building_sign_groups bsg JOIN sign_groups sg ON bsg.group_id=sg.id WHERE bsg.building_id=? ORDER BY sg.name''', conn, params=(building_id,))
    conn.close()
    st_opts = [{'label': f"{r.name} (${r.unit_price})", 'value': r.id} for r in st_df.itertuples()]
    table_rows = rows_df.to_dict('records')
    del_opts = [{'label': r['sign_name'], 'value': r['sign_name']} for r in table_rows]
    group_del_opts = [{'label': r.name, 'value': r.name} for r in grp_df.itertuples()] if not grp_df.empty else []
    meta = '' if b_df.empty else f"{b_df.iloc[0]['name']} - {b_df.iloc[0].get('description','')}"
    subtotal = sum(r['total'] for r in table_rows) if table_rows else 0
    group_count = 0 if grp_df.empty else grp_df.shape[0]
    summary = f"Subtotal: ${subtotal:,.2f} | Signs: {len(table_rows)} | Groups: {group_count}"
    return st_opts, table_rows, del_opts, group_del_opts, meta, summary

@app.callback(
    Output('bv-signs-table','data', allow_duplicate=True),
    Output('bv-delete-sign-dropdown','options', allow_duplicate=True),
    Output('bv-delete-group-dropdown','options', allow_duplicate=True),
    Output('bv-summary','children', allow_duplicate=True),
    Output('bv-feedback','children', allow_duplicate=True),
    Input('bv-add-sign-btn','n_clicks'),
    Input('bv-save-signs-btn','n_clicks'),
    Input('bv-delete-sign-btn','n_clicks'),
    Input('bv-delete-group-btn','n_clicks'),
    State('bv-building-dropdown','value'),
    State('bv-sign-type-dropdown','value'),
    State('bv-sign-qty-input','value'),
    State('bv-delete-sign-dropdown','value'),
    State('bv-delete-group-dropdown','value'),
    State('bv-signs-table','data'),
    prevent_initial_call=True
)
def bv_manage_signs(add_clicks, save_clicks, delete_clicks, delete_group_clicks, building_id, sign_type_id, qty, delete_name, delete_group_name, current_rows):
    triggered = [t['prop_id'].split('.')[0] for t in callback_context.triggered] if callback_context.triggered else []
    if not building_id:
        raise PreventUpdate
    msg = dash.no_update
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    if 'bv-add-sign-btn' in triggered and sign_type_id:
        q = max(1, int(qty or 1))
        cur.execute('SELECT id, quantity FROM building_signs WHERE building_id=? AND sign_type_id=?', (building_id, sign_type_id))
        ex = cur.fetchone()
        if ex:
            cur.execute('UPDATE building_signs SET quantity=? WHERE id=?', (q, ex[0]))
        else:
            cur.execute('INSERT INTO building_signs (building_id, sign_type_id, quantity) VALUES (?,?,?)', (building_id, sign_type_id, q))
        msg = 'Sign added/updated'
    elif 'bv-save-signs-btn' in triggered and current_rows:
        for r in current_rows:
            name = r.get('sign_name')
            q = max(0, int(r.get('quantity') or 0))
            cur.execute('SELECT id FROM sign_types WHERE name=?', (name,))
            st = cur.fetchone()
            if not st: continue
            cur.execute('SELECT id FROM building_signs WHERE building_id=? AND sign_type_id=?', (building_id, st[0]))
            ex = cur.fetchone()
            if ex:
                cur.execute('UPDATE building_signs SET quantity=? WHERE id=?', (q, ex[0]))
        msg = 'Quantities saved'
    elif 'bv-delete-sign-btn' in triggered and delete_name:
        cur.execute('SELECT id FROM sign_types WHERE name=?', (delete_name,))
        st = cur.fetchone()
        if st:
            cur.execute('DELETE FROM building_signs WHERE building_id=? AND sign_type_id=?', (building_id, st[0]))
            msg = 'Sign removed'
    elif 'bv-delete-group-btn' in triggered and delete_group_name:
        # Remove group assignment
        cur.execute('SELECT id FROM sign_groups WHERE name=?', (delete_group_name,))
        g = cur.fetchone()
        if g:
            cur.execute('DELETE FROM building_sign_groups WHERE building_id=? AND group_id=?', (building_id, g[0]))
            msg = 'Group removed'
    if msg is not dash.no_update:
        conn.commit()
    # Reload
    rows_df = pd.read_sql_query('''SELECT st.name as sign_name, bs.quantity, st.unit_price, (bs.quantity*st.unit_price) as total
                                   FROM building_signs bs JOIN sign_types st ON bs.sign_type_id=st.id
                                   WHERE bs.building_id=? ORDER BY st.name''', conn, params=(building_id,))
    grp_df = pd.read_sql_query('''SELECT sg.name, bsg.quantity FROM building_sign_groups bsg JOIN sign_groups sg ON bsg.group_id=sg.id WHERE bsg.building_id=? ORDER BY sg.name''', conn, params=(building_id,))
    conn.close()
    table_rows = rows_df.to_dict('records')
    del_opts = [{'label': r['sign_name'], 'value': r['sign_name']} for r in table_rows]
    group_del_opts = [{'label': r.name, 'value': r.name} for r in grp_df.itertuples()] if not grp_df.empty else []
    subtotal = sum(r['total'] for r in table_rows) if table_rows else 0
    summary = f"Subtotal: ${subtotal:,.2f} | Signs: {len(table_rows)} | Groups: {0 if grp_df.empty else grp_df.shape[0]}"
    return table_rows, del_opts, group_del_opts, summary, msg

@app.callback(
    Output('bv-building-dropdown','options', allow_duplicate=True),
    Output('bv-building-dropdown','value', allow_duplicate=True),
    Output('bv-building-meta','children', allow_duplicate=True),
    Input('bv-rename-building-btn','n_clicks'),
    State('bv-project-dropdown','value'),
    State('bv-building-dropdown','value'),
    State('bv-rename-building-input','value'),
    prevent_initial_call=True
)
def bv_rename_building(n_clicks, project_id, building_id, new_name):
    if not n_clicks:
        raise PreventUpdate
    if not (project_id and building_id and new_name and new_name.strip()):
        raise PreventUpdate
    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    cur.execute('SELECT 1 FROM buildings WHERE project_id=? AND LOWER(name)=LOWER(?) AND id<>?', (project_id, new_name.strip(), building_id))
    if cur.fetchone():
        conn.close()
        raise PreventUpdate
    cur.execute('UPDATE buildings SET name=?, last_modified=CURRENT_TIMESTAMP WHERE id=?', (new_name.strip(), building_id))
    conn.commit()
    bdf = pd.read_sql_query('SELECT id, name, description FROM buildings WHERE project_id=? ORDER BY name', conn, params=(project_id,))
    conn.close()
    opts = [{'label': r.name, 'value': r.id} for r in bdf.itertuples()]
    meta = ''
    for r in bdf.itertuples():
        if r.id == building_id:
            meta = f"{r.name} - {getattr(r,'description','')}"
            break
    return opts, building_id, meta

# ------------------ Health Endpoint ------------------ #
@app.server.route('/health')
def health():
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.execute('SELECT 1')
        conn.close()
        return {"status": "ok", "db": "reachable"}
    except Exception as e:
        return {"status": "error", "detail": str(e)}, 500

# ------------------ Port Selection Helper ------------------ #
def find_free_port(preferred: int) -> int:
    port = preferred
    for _ in range(15):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(('127.0.0.1', port))
                return port
            except OSError:
                port += 1
    return preferred  # fallback

if __name__ == "__main__":
    ensure_backup_dir()
    preferred = APP_PORT
    free_port = find_free_port(preferred)
    if free_port != preferred:
        print(f"[startup] Port {preferred} in use, switching to {free_port}")
    print("[startup] Starting Dash server ...")
    print(f"[startup] Database path: {DATABASE_PATH}")
    print(f"[startup] Open browser at: http://{APP_HOST}:{free_port}")
    if AUTO_BACKUP_INTERVAL_SEC > 0:
        print(f"[startup] Auto backup every {AUTO_BACKUP_INTERVAL_SEC}s -> {BACKUP_DIR}")
        import threading, time, shutil
        def _auto_backup_loop():
            while True:
                try:
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    target = BACKUP_DIR / f'sign_estimation_{ts}.db'
                    shutil.copy2(DATABASE_PATH, target)
                except Exception as e:
                    print(f"[backup][warn] {e}")
                time.sleep(AUTO_BACKUP_INTERVAL_SEC)
        threading.Thread(target=_auto_backup_loop, daemon=True).start()
    app.run(debug=DASH_DEBUG, port=free_port, host=APP_HOST, use_reloader=False)
